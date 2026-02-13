"""
Finance routes - payments, accounting endpoints з scope-based доступом.
"""
import logging
from datetime import datetime, date
from decimal import Decimal
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException, Request
from sqlalchemy.orm import Session

from core.database import get_db
from core.rbac import require_scope, Scope, filter_by_scope, get_user_scopes
from modules.auth.dependencies import get_current_user_db, role_required
from modules.auth.models import UserRole
from modules.finance.models import Transaction, PaymentMethod, PaymentStatus, Shipment, ShipmentMethod, ShipmentStatus
from modules.finance.schemas import ShipmentCreate, ShipmentRead, ShipmentUpdate
from modules.crm.models import Order
from modules.payment.models import PaymentSettings, PaymentProvider
from modules.payment.services.stripe_service import StripeService
from modules.payment.services.przelewy24_service import Przelewy24Service
from modules.payment.schemas import P24TransactionRegisterRequest
from pydantic import BaseModel
from typing import Optional, List
from uuid import UUID, uuid4
from decimal import Decimal
import os
from sqlalchemy.orm import joinedload
from sqlalchemy import and_, or_
import models

logger = logging.getLogger(__name__)

router = APIRouter(tags=["finance"])


@router.get("/revenue")
def get_revenue(
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT, UserRole.MANAGER])),
):
    """
    Отримати виручку.
    Доступ: менеджери, продажі, бухгалтер, адмін.
    """
    from core.rbac import get_user_scopes
    
    # Фільтруємо на основі scope
    # Якщо менеджер - показуємо всі продажі
    # Якщо sales-manager - показуємо тільки свої продажі
    
    user_scopes = get_user_scopes(user)
    
    if Scope.FINANCE_VIEW_REVENUE in user_scopes:
        # TODO: Реалізувати розрахунок виручки
        return {"revenue": 0, "currency": "UAH"}


@router.get("/profit")
def get_profit(
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT])),
):
    """
    Отримати прибуток (чистий прибуток).
    Доступ: ТІЛЬКИ бухгалтер та адмін.
    Менеджер НЕ бачить цей endpoint (403).
    """
    # TODO: Реалізувати розрахунок прибутку
    # Прибуток = Виручка - Витрати
    return {"profit": 0, "currency": "UAH", "note": "Only accountant and admin can see this"}


@router.get("/costs")
def get_costs(
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT])),
):
    """
    Отримати витрати.
    Доступ: бухгалтер, адмін.
    """
    # TODO: Реалізувати розрахунок витрат
    return {"costs": 0, "currency": "UAH"}


@router.get("/payments")
def get_payments(
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT, UserRole.MANAGER])),
):
    """
    Отримати список платежів (транзакцій).
    Менеджери бачать тільки свої платежі.
    Бухгалтер та адмін бачать всі.
    """
    user_scopes = get_user_scopes(user)
    
    # Запит транзакцій з завантаженням order та client
    query = (
        db.query(Transaction)
        .options(
            joinedload(Transaction.order).joinedload(Order.client)
        )
        .order_by(Transaction.payment_date.desc())
    )
    
    # Фільтрація на основі scope
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        # Тільки свої платежі (де order.manager_id == user.id)
        query = query.join(Order).filter(Order.manager_id == user.id)
    
    transactions = query.all()
    
    # Формуємо відповідь
    payments = []
    for idx, trans in enumerate(transactions, 1):
        order = trans.order
        client = order.client if order else None
        buyer_name = client.full_name if client else "N/A"
        
        payment_data = {
            "id": str(trans.id),
            "lp": idx,  # Порядковий номер
            "order_number": order.order_number if order else "N/A",
            "service_date": trans.service_date.isoformat() if trans.service_date else None,
            "buyer_name": buyer_name,
            "amount_gross": float(trans.amount_gross),
            "payment_date": trans.payment_date.isoformat() if trans.payment_date else None,
            "posting_date": trans.posting_date.isoformat() if trans.posting_date else None,
            "payment_method": trans.payment_method.value if isinstance(trans.payment_method, PaymentMethod) else trans.payment_method,
            "receipt_number": trans.receipt_number,
            "notes": trans.notes,
            # Stripe fields
            "stripe_payment_intent_id": trans.stripe_payment_intent_id,
            "stripe_session_id": trans.stripe_session_id,
            "stripe_customer_email": trans.stripe_customer_email,
            "currency": trans.currency or "PLN",
            "stripe_fee": float(trans.stripe_fee) if trans.stripe_fee else None,
            "net_amount": float(trans.net_amount) if trans.net_amount else None,
            "card_brand": trans.card_brand,
            "card_last4": trans.card_last4,
            "stripe_receipt_url": trans.stripe_receipt_url,
            "payment_status": trans.payment_status.value if isinstance(trans.payment_status, PaymentStatus) else trans.payment_status,
            "stripe_payment_link_id": trans.stripe_payment_link_id,
        }
        payments.append(payment_data)
    
    return {"payments": payments}


@router.get("/payments/export")
def export_payments_excel(
    db: Session = Depends(get_db),
    user: models.User = Depends(get_current_user_db),
):
    """
    Експортувати платежі в Excel.
    Менеджери бачать тільки свої платежі.
    Бухгалтер та адмін бачать всі.
    """
    from fastapi.responses import Response
    from datetime import datetime
    from io import BytesIO
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is required for Excel export")
    
    user_scopes = get_user_scopes(user)
    
    # Запит транзакцій з завантаженням order та client
    query = (
        db.query(Transaction)
        .options(
            joinedload(Transaction.order).joinedload(Order.client)
        )
        .order_by(Transaction.payment_date.desc())
    )
    
    # Фільтрація на основі scope
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        # Тільки свої платежі (де order.manager_id == user.id)
        query = query.join(Order).filter(Order.manager_id == user.id)
    
    transactions = query.all()
    
    # Створюємо Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Płatności"
    
    # Стилі
    header_font = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Кольори для заголовків (згідно з зображенням)
    header_colors = [
        "90EE90",  # LP - зелений
        "87CEEB",  # Numer zlecenia - синій
        "DDA0DD",  # Data wykonania usługi - фіолетовий
        "FFD700",  # Nabywca - жовтий
        "FFA500",  # Kwota płatności brutto - помаранчевий
        "FFB6C1",  # Data płatności - світло-червоний
        "87CEEB",  # Data nabicia na KF - синій
        "FFC0CB",  # Sposób płatności - рожевий
        "FFB6C1",  # Numer dowodu sprzedaży - світло-червоний
        "D3D3D3",  # Uwagi - сірий
    ]
    
    # Заголовки
    headers = [
        "LP",
        "Numer zlecenia",
        "Data wykonania usługi",
        "Nabywca (Imię i nazwisko)",
        "Kwota płatności brutto",
        "Data płatności",
        "Data nabicia na KF",
        "Sposób płatności",
        "Numer dowodu sprzedaży",
        "Uwagi",
    ]
    
    for col_idx, (header, color) in enumerate(zip(headers, header_colors), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Дані
    for row_idx, trans in enumerate(transactions, 2):
        order = trans.order
        client = order.client if order else None
        buyer_name = client.full_name if client else "N/A"
        
        # Форматування способу оплати
        payment_method_map = {
            "transfer": "Przelew",
            "card": "Karta",
            "blik": "BLIK",
            "cash": "Gotówka",
        }
        payment_method_display = payment_method_map.get(
            trans.payment_method.value if isinstance(trans.payment_method, PaymentMethod) else trans.payment_method,
            str(trans.payment_method)
        )
        
        # Форматування дат
        def format_date(d):
            if d:
                return d.strftime("%d.%m.%Y")
            return ""
        
        row_data = [
            row_idx - 1,  # LP
            order.order_number if order else "N/A",
            format_date(trans.service_date),
            buyer_name,
            float(trans.amount_gross),
            format_date(trans.payment_date),
            format_date(trans.posting_date),
            payment_method_display,
            trans.receipt_number,
            trans.notes or "",
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 5:  # Kwota płatności brutto
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx in [3, 6, 7]:  # Дати
                cell.alignment = Alignment(horizontal="center")
    
    # Налаштування ширини колонок
    column_widths = [5, 25, 18, 25, 20, 15, 18, 18, 25, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Зберігаємо в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Формуємо назву файлу
    filename = f"platnosci_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get("/accounting")
def get_accounting(
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT])),
):
    """
    Отримати бухгалтерські звіти.
    Доступ: ТІЛЬКИ бухгалтер та адмін.
    """
    # TODO: Реалізувати бухгалтерські звіти
    return {"reports": []}


@router.post("/webhooks/stripe")
async def stripe_webhook(
    request: Request,
    db: Session = Depends(get_db),
):
    """
    Stripe webhook handler для автоматичного створення/оновлення Finance Transaction.
    Обробляє події: checkout.session.completed, payment_intent.succeeded, charge.refunded.
    """
    settings = db.query(PaymentSettings).first()
    if not settings or not settings.stripe_enabled or not settings.stripe_webhook_secret:
        raise HTTPException(status_code=400, detail="Stripe not configured")
    
    payload = await request.body()
    sig_header = request.headers.get("stripe-signature")
    
    if not sig_header:
        raise HTTPException(status_code=400, detail="Missing signature")
    
    service = StripeService(settings)
    
    try:
        event = service.verify_webhook_signature(payload, sig_header)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid signature")
    
    event_type = event["type"]
    data = event["data"]["object"]
    
    logger.info(f"Finance Stripe webhook received: {event_type}")
    
    try:
        if event_type == "checkout.session.completed":
            await _handle_checkout_session_completed(data, db, service)
        elif event_type == "payment_intent.succeeded":
            await _handle_payment_intent_succeeded(data, db, service)
        elif event_type == "charge.refunded":
            await _handle_charge_refunded(data, db)
        else:
            logger.info(f"Unhandled Stripe event type: {event_type}")
        
        db.commit()
        return {"received": True}
    except Exception as e:
        logger.error(f"Error processing Stripe webhook: {e}", exc_info=True)
        db.rollback()
        # Return 200 OK to Stripe to avoid retries, but log the error
        return {"received": True, "error": str(e)}


async def _handle_checkout_session_completed(data: dict, db: Session, service: StripeService):
    """Handle checkout.session.completed event."""
    session_id = data.get("id")
    metadata = data.get("metadata", {})
    order_id_str = metadata.get("order_id")
    payment_link_id = data.get("payment_link")
    
    if not order_id_str:
        logger.warning(f"Stripe webhook: missing order_id in metadata for session {session_id}")
        return
    
    try:
        order_id = UUID(order_id_str)
    except (ValueError, TypeError):
        logger.warning(f"Stripe webhook: invalid order_id in metadata: {order_id_str}")
        return
    
    # Get order
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        logger.error(f"Stripe webhook: order not found: {order_id}")
        return
    
    # Get payment intent details
    payment_intent_id = data.get("payment_intent")
    if not payment_intent_id:
        logger.warning(f"Stripe webhook: missing payment_intent in checkout session {session_id}")
        return
    
    # Retrieve payment intent from Stripe
    import stripe
    stripe.api_key = service.secret_key
    try:
        payment_intent = stripe.PaymentIntent.retrieve(payment_intent_id)
        charge = payment_intent.charges.data[0] if payment_intent.charges and payment_intent.charges.data else None
    except Exception as e:
        logger.error(f"Stripe webhook: failed to retrieve payment intent {payment_intent_id}: {e}")
        return
    
    # Extract customer email
    customer_email = data.get("customer_email") or (payment_intent.get("receipt_email") if payment_intent else None)
    
    # Extract amount and currency
    amount_total = Decimal(str(data.get("amount_total", 0))) / 100  # Stripe uses cents
    currency = data.get("currency", "pln").upper()
    
    # Calculate Stripe fee and net amount
    stripe_fee = None
    net_amount = None
    if charge:
        balance_transaction_id = charge.get("balance_transaction")
        if balance_transaction_id:
            try:
                balance_transaction = stripe.BalanceTransaction.retrieve(balance_transaction_id)
                stripe_fee = Decimal(str(balance_transaction.fee)) / 100
                net_amount = amount_total - stripe_fee
            except Exception as e:
                logger.warning(f"Stripe webhook: failed to retrieve balance transaction: {e}")
    
    # Extract card details
    card_brand = None
    card_last4 = None
    if charge and charge.get("payment_method_details"):
        card = charge.payment_method_details.get("card")
        if card:
            card_brand = card.get("brand", "").title()  # visa -> Visa
            card_last4 = card.get("last4")
    
    # Get receipt URL
    receipt_url = charge.get("receipt_url") if charge else None
    
    # Check if transaction already exists
    existing_transaction = db.query(Transaction).filter(
        Transaction.stripe_payment_intent_id == payment_intent_id
    ).first()
    
    if existing_transaction:
        # Update existing transaction
        existing_transaction.stripe_session_id = session_id
        existing_transaction.stripe_customer_email = customer_email
        existing_transaction.currency = currency
        existing_transaction.stripe_fee = stripe_fee
        existing_transaction.net_amount = net_amount
        existing_transaction.card_brand = card_brand
        existing_transaction.card_last4 = card_last4
        existing_transaction.stripe_receipt_url = receipt_url
        existing_transaction.payment_status = PaymentStatus.SUCCEEDED
        existing_transaction.stripe_payment_link_id = payment_link_id
        existing_transaction.amount_gross = amount_total
        logger.info(f"Updated finance transaction for payment intent {payment_intent_id}")
    else:
        # Create new transaction
        payment_method = PaymentMethod.CARD if card_brand else PaymentMethod.TRANSFER
        
        new_transaction = Transaction(
            order_id=order_id,
            amount_gross=amount_total,
            payment_date=date.today(),
            service_date=date.today(),
            receipt_number=f"STRIPE-{payment_intent_id[:8]}",
            payment_method=payment_method,
            notes=f"Automatic payment via Stripe Payment Link",
            stripe_payment_intent_id=payment_intent_id,
            stripe_session_id=session_id,
            stripe_customer_email=customer_email,
            currency=currency,
            stripe_fee=stripe_fee,
            net_amount=net_amount,
            card_brand=card_brand,
            card_last4=card_last4,
            stripe_receipt_url=receipt_url,
            payment_status=PaymentStatus.SUCCEEDED,
            stripe_payment_link_id=payment_link_id,
        )
        db.add(new_transaction)
        logger.info(f"Created finance transaction for payment intent {payment_intent_id}")
    
    # Update order status
    order.status = "oplacone"
    order.payment_method = "card" if card_brand else "transfer"
    
    logger.info(f"Stripe webhook: successfully processed checkout.session.completed for order {order_id}")


async def _handle_payment_intent_succeeded(data: dict, db: Session, service: StripeService):
    """Handle payment_intent.succeeded event."""
    payment_intent_id = data.get("id")
    metadata = data.get("metadata", {})
    order_id_str = metadata.get("order_id")
    
    if not order_id_str:
        logger.warning(f"Stripe webhook: missing order_id in metadata for payment_intent {payment_intent_id}")
        return
    
    try:
        order_id = UUID(order_id_str)
    except (ValueError, TypeError):
        logger.warning(f"Stripe webhook: invalid order_id in metadata: {order_id_str}")
        return
    
    # Get order
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        logger.error(f"Stripe webhook: order not found: {order_id}")
        return
    
    # Extract amount and currency
    amount_total = Decimal(str(data.get("amount", 0))) / 100
    currency = data.get("currency", "pln").upper()
    
    # Get charge details
    charges = data.get("charges", {}).get("data", [])
    charge = charges[0] if charges else None
    
    # Extract customer email
    customer_email = data.get("receipt_email")
    
    # Calculate Stripe fee and net amount
    stripe_fee = None
    net_amount = None
    if charge:
        balance_transaction_id = charge.get("balance_transaction")
        if balance_transaction_id:
            import stripe
            stripe.api_key = service.secret_key
            try:
                balance_transaction = stripe.BalanceTransaction.retrieve(balance_transaction_id)
                stripe_fee = Decimal(str(balance_transaction.fee)) / 100
                net_amount = amount_total - stripe_fee
            except Exception as e:
                logger.warning(f"Stripe webhook: failed to retrieve balance transaction: {e}")
    
    # Extract card details
    card_brand = None
    card_last4 = None
    if charge and charge.get("payment_method_details"):
        card = charge.payment_method_details.get("card")
        if card:
            card_brand = card.get("brand", "").title()
            card_last4 = card.get("last4")
    
    # Get receipt URL
    receipt_url = charge.get("receipt_url") if charge else None
    
    # Check if transaction already exists
    existing_transaction = db.query(Transaction).filter(
        Transaction.stripe_payment_intent_id == payment_intent_id
    ).first()
    
    if existing_transaction:
        # Update existing transaction
        existing_transaction.stripe_customer_email = customer_email
        existing_transaction.currency = currency
        existing_transaction.stripe_fee = stripe_fee
        existing_transaction.net_amount = net_amount
        existing_transaction.card_brand = card_brand
        existing_transaction.card_last4 = card_last4
        existing_transaction.stripe_receipt_url = receipt_url
        existing_transaction.payment_status = PaymentStatus.SUCCEEDED
        existing_transaction.amount_gross = amount_total
        logger.info(f"Updated finance transaction for payment intent {payment_intent_id}")
    else:
        # Create new transaction
        payment_method = PaymentMethod.CARD if card_brand else PaymentMethod.TRANSFER
        
        new_transaction = Transaction(
            order_id=order_id,
            amount_gross=amount_total,
            payment_date=date.today(),
            service_date=date.today(),
            receipt_number=f"STRIPE-{payment_intent_id[:8]}",
            payment_method=payment_method,
            notes=f"Automatic payment via Stripe",
            stripe_payment_intent_id=payment_intent_id,
            stripe_customer_email=customer_email,
            currency=currency,
            stripe_fee=stripe_fee,
            net_amount=net_amount,
            card_brand=card_brand,
            card_last4=card_last4,
            stripe_receipt_url=receipt_url,
            payment_status=PaymentStatus.SUCCEEDED,
        )
        db.add(new_transaction)
        logger.info(f"Created finance transaction for payment intent {payment_intent_id}")
    
    # Update order status
    order.status = "oplacone"
    order.payment_method = "card" if card_brand else "transfer"
    
    logger.info(f"Stripe webhook: successfully processed payment_intent.succeeded for order {order_id}")


async def _handle_charge_refunded(data: dict, db: Session):
    """Handle charge.refunded event."""
    charge_id = data.get("id")
    payment_intent_id = data.get("payment_intent")
    
    if not payment_intent_id:
        logger.warning(f"Stripe webhook: missing payment_intent in charge {charge_id}")
        return
    
    # Find transaction by payment_intent_id
    transaction = db.query(Transaction).filter(
        Transaction.stripe_payment_intent_id == payment_intent_id
    ).first()
    
    if not transaction:
        logger.warning(f"Stripe webhook: transaction not found for payment_intent {payment_intent_id}")
        return
    
    # Update transaction status
    transaction.payment_status = PaymentStatus.REFUNDED
    transaction.notes = (transaction.notes or "") + f"\nRefunded on {datetime.utcnow().isoformat()}"
    
    # Update order status back to unpaid
    order = db.query(Order).filter(Order.id == transaction.order_id).first()
    if order:
        order.status = "do_wykonania"
    
    logger.info(f"Stripe webhook: marked transaction {transaction.id} as refunded")


class PaymentLinkRequest(BaseModel):
    amount: Optional[float] = None
    currency: str = "pln"


@router.post("/payment-link/{order_id}")
async def create_payment_link(
    order_id: UUID,
    request: PaymentLinkRequest,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.MANAGER])),
):
    """
    Create payment link for order using active payment provider (Stripe or Przelewy24).
    This endpoint belongs to finance module as it's a financial operation.
    """
    amount = request.amount
    currency = request.currency or "PLN"
    
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # If amount not provided, calculate from order or use default
    if not amount:
        if order.price_brutto:
            amount = float(order.price_brutto)
            logger.info(f"Payment link: using order.price_brutto = {amount} for order {order_id}")
        else:
            amount = 100.0
            logger.warning(f"Payment link: order.price_brutto is None, using default 100.0 for order {order_id}")
    else:
        logger.info(f"Payment link: using provided amount = {amount} for order {order_id}")
    
    logger.info(f"Payment link: final amount = {amount} for order {order_id} (order.price_brutto = {order.price_brutto})")
    
    # Get payment settings
    payment_settings = db.query(PaymentSettings).first()
    if not payment_settings:
        raise HTTPException(status_code=500, detail="Payment settings not configured")
    
    # Determine active payment provider
    active_provider = payment_settings.active_payment_provider
    
    # Convert string to Enum if needed
    if isinstance(active_provider, str):
        try:
            active_provider = PaymentProvider(active_provider.lower())
        except ValueError:
            active_provider = None
    
    # If not set, auto-select based on enabled statuses
    if not active_provider:
        if payment_settings.stripe_enabled and payment_settings.stripe_secret_key:
            active_provider = PaymentProvider.STRIPE
        elif payment_settings.przelewy24_enabled and payment_settings.przelewy24_merchant_id:
            active_provider = PaymentProvider.PRZELEWY24
        else:
            raise HTTPException(
                status_code=500, 
                detail="No active payment provider configured. Please configure Stripe or Przelewy24 in settings."
            )
    
    # Create payment link based on provider
    try:
        if active_provider == PaymentProvider.STRIPE:
            # Stripe payment link
            if not payment_settings.stripe_enabled or not payment_settings.stripe_secret_key:
                raise HTTPException(status_code=500, detail="Stripe is not configured or enabled")
            
            import stripe
            stripe.api_key = payment_settings.stripe_secret_key
            
            # Create Price
            price = stripe.Price.create(
                unit_amount=int(amount * 100),  # Stripe works in cents
                currency=currency.lower(),
                product_data={
                    "name": f"Замовлення {order.order_number}",
                },
            )

            # Create Payment Link with order_id in metadata for webhook
            payment_link_obj = stripe.PaymentLink.create(
                line_items=[
                    {
                        "price": price.id,
                        "quantity": 1,
                    },
                ],
                metadata={
                    "order_id": str(order.id),
                },
                payment_intent_data={
                    "metadata": {
                        "order_id": str(order.id),
                    }
                },
                after_completion={
                    "type": "redirect",
                    "redirect": {
                        "url": f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/orders/{order.id}/success"
                    }
                }
            )

            payment_url = payment_link_obj.url
            payment_link_id = payment_link_obj.id
            
        elif active_provider == PaymentProvider.PRZELEWY24:
            # Przelewy24 payment link
            if not payment_settings.przelewy24_enabled or not payment_settings.przelewy24_merchant_id:
                raise HTTPException(status_code=500, detail="Przelewy24 is not configured or enabled")
            
            p24_service = Przelewy24Service(payment_settings)
            
            # Get customer email
            customer_email = order.client.email if order.client and order.client.email else "customer@example.com"
            customer_name = order.client.full_name if order.client else "Клієнт"
            
            # Create session_id
            session_id = str(uuid4())
            
            # URLs
            frontend_url = os.getenv('FRONTEND_URL', 'http://localhost:5173')
            backend_url = os.getenv('BACKEND_URL', 'http://localhost:8000')
            return_url = f"{frontend_url}/orders/{order.id}/success"
            status_url = f"{backend_url}/api/v1/payment/webhooks/przelewy24"
            
            # Register transaction in Przelewy24
            p24_request = P24TransactionRegisterRequest(
                order_id=order.id,
                amount=Decimal(str(amount)),
                currency=currency,
                customer_email=customer_email,
                customer_name=customer_name,
                description=f"Замовлення {order.order_number}",
            )
            
            result = await p24_service.register_transaction(
                request=p24_request,
                session_id=session_id,
                return_url=return_url,
                status_url=status_url,
            )
            
            payment_url = result.payment_url
            payment_link_id = None  # P24 doesn't have payment link ID
        else:
            raise HTTPException(status_code=500, detail=f"Unsupported payment provider: {active_provider}")

        # Convert provider to string (handle both Enum and str)
        provider_str = active_provider.value if hasattr(active_provider, 'value') else str(active_provider)
        
        return {
            "payment_link": payment_url,
            "order_id": str(order.id),
            "amount": amount,
            "currency": currency,
            "provider": provider_str,
            "payment_link_id": payment_link_id,  # Stripe Payment Link ID
        }
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"Payment library not installed: {str(e)}")
    except Exception as e:
        logger.error(f"Payment error ({active_provider}): {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create payment link: {str(e)}")


# Shipments endpoints
@router.get("/shipments", response_model=List[ShipmentRead])
def get_shipments(
    status: Optional[str] = None,
    method: Optional[str] = None,
    order_id: Optional[UUID] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT, UserRole.MANAGER])),
):
    """
    Отримати список відправок з фільтрами.
    Менеджери бачать тільки свої відправки.
    Бухгалтер та адмін бачать всі.
    """
    user_scopes = get_user_scopes(user)
    
    # Запит відправок з завантаженням order
    query = (
        db.query(Shipment)
        .options(joinedload(Shipment.order))
        .order_by(Shipment.created_at.desc())
    )
    
    # Фільтрація на основі scope
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        # Тільки свої відправки (де order.manager_id == user.id)
        query = query.join(Order).filter(Order.manager_id == user.id)
    
    # Фільтри
    if order_id:
        query = query.filter(Shipment.order_id == order_id)
    if status:
        query = query.filter(Shipment.status == status)
    if method:
        query = query.filter(Shipment.method == method)
    
    # Pagination
    shipments = query.offset(skip).limit(limit).all()
    
    # Формуємо відповідь
    result = []
    for shipment in shipments:
        shipment_data = ShipmentRead.model_validate(shipment)
        shipment_data.order_number = shipment.order.order_number if shipment.order else None
        result.append(shipment_data)
    
    return result


@router.get("/shipments/{shipment_id}", response_model=ShipmentRead)
def get_shipment(
    shipment_id: UUID,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.ACCOUNTANT, UserRole.MANAGER])),
):
    """Отримати деталі відправки."""
    shipment = (
        db.query(Shipment)
        .options(joinedload(Shipment.order))
        .filter(Shipment.id == shipment_id)
        .first()
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    # Перевірка доступу
    user_scopes = get_user_scopes(user)
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        if shipment.order.manager_id != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    shipment_data = ShipmentRead.model_validate(shipment)
    shipment_data.order_number = shipment.order.order_number if shipment.order else None
    return shipment_data


@router.post("/shipments", response_model=ShipmentRead, status_code=201)
async def create_shipment(
    shipment_data: ShipmentCreate,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.MANAGER])),
):
    """Створити відправку. Можна одразу створити InPost shipment через API."""
    # Перевіряємо замовлення з явним завантаженням клієнта
    order = db.query(Order).options(joinedload(Order.client)).filter(Order.id == shipment_data.order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Генеруємо tracking_url для InPost
    tracking_url = None
    if shipment_data.method in [ShipmentMethod.INPOST_LOCKER, ShipmentMethod.INPOST_COURIER]:
        if shipment_data.tracking_number:
            tracking_url = f"https://inpost.pl/sledzenie-przesylek?number={shipment_data.tracking_number}"
    
    # Створюємо відправку
    shipment = Shipment(
        order_id=shipment_data.order_id,
        method=shipment_data.method,
        tracking_number=shipment_data.tracking_number,
        tracking_url=tracking_url,
        status=ShipmentStatus.CREATED,
        paczkomat_code=shipment_data.paczkomat_code,
        delivery_address=shipment_data.delivery_address,
        recipient_name=shipment_data.recipient_name,
        recipient_email=shipment_data.recipient_email,
        recipient_phone=shipment_data.recipient_phone,
        shipping_cost=shipment_data.shipping_cost or Decimal("13.99"),  # Default для InPost
    )
    
    db.add(shipment)
    db.flush()
    
    # Якщо потрібно створити через InPost API
    if shipment_data.create_inpost_shipment and shipment_data.method in [ShipmentMethod.INPOST_LOCKER, ShipmentMethod.INPOST_COURIER]:
        try:
            from modules.postal_services.service import InPostService
            from modules.postal_services.schemas import CreateShipmentRequest, ReceiverInfo, AddressInfo
            from modules.postal_services.models import DeliveryType
            
            inpost_service = InPostService(db)
            
            # Створюємо запит для InPost
            delivery_type = DeliveryType.PARCEL_LOCKER if shipment_data.method == ShipmentMethod.INPOST_LOCKER else DeliveryType.COURIER
            
            # Перевіряємо, чи є існуючий Shipment запис для цього замовлення
            existing_shipment = db.query(Shipment).filter(
                Shipment.order_id == order.id,
                Shipment.method.in_([ShipmentMethod.INPOST_LOCKER, ShipmentMethod.INPOST_COURIER])
            ).order_by(Shipment.created_at.desc()).first()
            
            # Отримуємо дані отримувача
            # Перевіряємо, чи завантажений клієнт
            if not order.client:
                raise HTTPException(
                    status_code=400,
                    detail="Клієнт не знайдений для цього замовлення."
                )
            
            # Отримуємо email та телефон, перевіряючи на порожні рядки
            client_email = order.client.email.strip() if order.client.email and order.client.email.strip() else None
            client_phone = order.client.phone.strip() if order.client.phone and order.client.phone.strip() else None
            client_name = order.client.full_name if order.client.full_name else "Клієнт"
            
            # Використовуємо дані з існуючого Shipment, якщо вони є, інакше з shipment_data або з клієнта
            recipient_email = shipment_data.recipient_email or (existing_shipment.recipient_email if existing_shipment and existing_shipment.recipient_email else None) or client_email
            recipient_phone = shipment_data.recipient_phone or (existing_shipment.recipient_phone if existing_shipment and existing_shipment.recipient_phone else None) or client_phone
            recipient_name = shipment_data.recipient_name or (existing_shipment.recipient_name if existing_shipment and existing_shipment.recipient_name else None) or client_name
            
            # Використовуємо пачкомат та адресу з існуючого Shipment, якщо вони є
            # Перевіряємо на None та порожні рядки
            paczkomat_code = None
            if shipment_data.paczkomat_code and shipment_data.paczkomat_code.strip():
                paczkomat_code = shipment_data.paczkomat_code
            elif existing_shipment and existing_shipment.paczkomat_code and existing_shipment.paczkomat_code.strip():
                paczkomat_code = existing_shipment.paczkomat_code
            
            delivery_address = shipment_data.delivery_address or (existing_shipment.delivery_address if existing_shipment else None)
            
            # Логування для діагностики
            logger.info(f"Creating shipment for order {order.id}: client_id={order.client_id}, client_email={client_email}, client_phone={client_phone}")
            logger.info(f"Final recipient data: email={recipient_email}, phone={recipient_phone}, name={recipient_name}")
            logger.info(f"Paczkomat code: shipment_data.paczkomat_code='{shipment_data.paczkomat_code}', existing_shipment.paczkomat_code='{existing_shipment.paczkomat_code if existing_shipment else None}', final='{paczkomat_code}'")
            print(f"[InPost] Paczkomat code: from shipment_data='{shipment_data.paczkomat_code}', from existing='{existing_shipment.paczkomat_code if existing_shipment else None}', final='{paczkomat_code}'")
            
            # Перевіряємо, чи є email та телефон (перевіряємо на None та порожні рядки)
            if not recipient_email or not recipient_email.strip():
                raise HTTPException(
                    status_code=400,
                    detail=f"Email клієнта обов'язковий для створення відправлення InPost. Додайте email до картки клієнта. (Client ID: {order.client_id}, Current email: '{client_email}')"
                )
            if not recipient_phone or not recipient_phone.strip():
                raise HTTPException(
                    status_code=400,
                    detail=f"Телефон клієнта обов'язковий для створення відправлення InPost. Додайте телефон до картки клієнта. (Client ID: {order.client_id}, Current phone: '{client_phone}')"
                )
            
            # Створюємо об'єкт ReceiverInfo
            receiver = ReceiverInfo(
                email=recipient_email,
                phone=recipient_phone,
                name=recipient_name,
            )
            
            # Обробляємо адресу для кур'єрської доставки
            courier_address_obj = None
            if shipment_data.method == ShipmentMethod.INPOST_COURIER:
                if not shipment_data.delivery_address:
                    raise HTTPException(
                        status_code=400,
                        detail="Адреса доставки обов'язкова для кур'єрської доставки InPost."
                    )
                # Якщо delivery_address - це рядок, спробуємо розпарсити його
                # Формат може бути: "XX-XXX Місто, Вулиця Номер/Квартира" або просто текст
                # Для InPost API потрібна структурована адреса, тому спробуємо розпарсити
                import re
                address_str = shipment_data.delivery_address
                
                # Спроба розпарсити польський формат адреси: "XX-XXX Місто, Вулиця Номер/Квартира"
                # Або "XX-XXX Місто, Вулиця Номер"
                post_code_match = re.search(r'(\d{2}-\d{3})', address_str)
                if post_code_match:
                    post_code = post_code_match.group(1)
                    # Видаляємо поштовий індекс з рядка для подальшого парсингу
                    remaining = address_str.replace(post_code, '').strip()
                    
                    # Спроба знайти місто та вулицю
                    # Формат може бути: "Місто, Вулиця Номер" або "Місто Вулиця Номер"
                    parts = [p.strip() for p in remaining.split(',') if p.strip()]
                    if len(parts) >= 2:
                        city = parts[0]
                        street_part = parts[1]
                    elif len(parts) == 1:
                        # Спроба розділити по пробілах
                        words = parts[0].split()
                        if len(words) >= 3:
                            city = ' '.join(words[:-2])
                            street_part = ' '.join(words[-2:])
                        else:
                            city = parts[0]
                            street_part = ""
                    else:
                        city = ""
                        street_part = remaining
                    
                    # Парсинг вулиці та номера будинку
                    street_match = re.match(r'^(.+?)\s+(\d+[A-Za-z]?)(?:/(\d+))?$', street_part.strip())
                    if street_match:
                        street = street_match.group(1).strip()
                        building_number = street_match.group(2)
                        flat_number = street_match.group(3) if street_match.group(3) else None
                    else:
                        # Якщо не вдалося розпарсити, використовуємо весь рядок як вулицю
                        street = street_part.strip() or remaining
                        building_number = "1"
                        flat_number = None
                    
                    if city and street and building_number:
                        courier_address_obj = AddressInfo(
                            street=street,
                            building_number=building_number,
                            flat_number=flat_number,
                            city=city,
                            post_code=post_code,
                            country="PL"
                        )
                
                # Якщо не вдалося розпарсити адресу, показуємо помилку
                if not courier_address_obj:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Не вдалося розпарсити адресу доставки. Потрібен формат: 'XX-XXX Місто, Вулиця Номер/Квартира'. Отримано: {shipment_data.delivery_address}"
                    )
            
            # Normalize parcel locker code if provided
            normalized_paczkomat_code = None
            if shipment_data.method == ShipmentMethod.INPOST_LOCKER:
                if not paczkomat_code or not paczkomat_code.strip():
                    error_msg = f"Parcel locker code is required for INPOST_LOCKER delivery method. Order ID: {order.id}"
                    logger.error(error_msg)
                    print(f"[InPost] ERROR: {error_msg}")
                    raise HTTPException(status_code=400, detail=error_msg)
                
                normalized_paczkomat_code = paczkomat_code.strip().upper()
                logger.info(f"InPost: Normalizing parcel locker code from finance router: '{paczkomat_code}' -> '{normalized_paczkomat_code}'")
                print(f"[InPost] Normalized paczkomat code: '{paczkomat_code}' -> '{normalized_paczkomat_code}'")
            
            create_request = CreateShipmentRequest(
                order_id=order.id,
                delivery_type=delivery_type,
                parcel_locker_code=normalized_paczkomat_code,
                receiver=receiver,
                courier_address=courier_address_obj,
            )
            
            # Створюємо відправку в InPost
            inpost_shipment = await inpost_service.create_shipment(create_request)
            
            # Оновлюємо наш shipment з даними з InPost
            shipment.tracking_number = inpost_shipment.tracking_number
            shipment.tracking_url = inpost_shipment.tracking_url
            shipment.label_url = inpost_shipment.label_url
            shipment.inpost_shipment_id = inpost_shipment.shipment_id
            shipment.status = ShipmentStatus.LABEL_PRINTED if inpost_shipment.label_url else ShipmentStatus.CREATED
            
        except Exception as e:
            logger.error(f"Failed to create InPost shipment: {e}", exc_info=True)
            # Продовжуємо без InPost - відправка створена локально
    
    db.commit()
    db.refresh(shipment)
    
    shipment_read = ShipmentRead.model_validate(shipment)
    shipment_read.order_number = order.order_number
    return shipment_read


@router.patch("/shipments/{shipment_id}", response_model=ShipmentRead)
def update_shipment(
    shipment_id: UUID,
    shipment_update: ShipmentUpdate,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.MANAGER])),
):
    """Оновити відправку."""
    shipment = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    # Перевірка доступу
    user_scopes = get_user_scopes(user)
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        if shipment.order.manager_id != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    # Оновлюємо поля
    if shipment_update.tracking_number is not None:
        shipment.tracking_number = shipment_update.tracking_number
        # Auto-generate tracking_url для InPost
        if shipment.method in [ShipmentMethod.INPOST_LOCKER, ShipmentMethod.INPOST_COURIER] and shipment.tracking_number:
            shipment.tracking_url = f"https://inpost.pl/sledzenie-przesylek?number={shipment.tracking_number}"
    
    if shipment_update.status is not None:
        shipment.status = shipment_update.status
    
    if shipment_update.paczkomat_code is not None:
        shipment.paczkomat_code = shipment_update.paczkomat_code
    
    if shipment_update.delivery_address is not None:
        shipment.delivery_address = shipment_update.delivery_address
    
    if shipment_update.recipient_name is not None:
        shipment.recipient_name = shipment_update.recipient_name
    
    if shipment_update.recipient_email is not None:
        shipment.recipient_email = shipment_update.recipient_email
    
    if shipment_update.recipient_phone is not None:
        shipment.recipient_phone = shipment_update.recipient_phone
    
    if shipment_update.shipping_cost is not None:
        shipment.shipping_cost = shipment_update.shipping_cost
    
    if shipment_update.label_url is not None:
        shipment.label_url = shipment_update.label_url
    
    if shipment_update.shipped_at is not None:
        shipment.shipped_at = shipment_update.shipped_at
    
    if shipment_update.delivered_at is not None:
        shipment.delivered_at = shipment_update.delivered_at
    
    if shipment_update.inpost_shipment_id is not None:
        shipment.inpost_shipment_id = shipment_update.inpost_shipment_id
    
    db.commit()
    db.refresh(shipment)
    
    shipment_read = ShipmentRead.model_validate(shipment)
    shipment_read.order_number = shipment.order.order_number if shipment.order else None
    return shipment_read


@router.post("/shipments/{shipment_id}/track")
async def track_shipment(
    shipment_id: UUID,
    db: Session = Depends(get_db),
    user: models.User = Depends(role_required([UserRole.OWNER, UserRole.MANAGER])),
):
    """Оновити статус відправки з InPost API."""
    shipment = (
        db.query(Shipment)
        .options(joinedload(Shipment.order))
        .filter(Shipment.id == shipment_id)
        .first()
    )
    
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    # Перевірка доступу
    user_scopes = get_user_scopes(user)
    if Scope.CRM_VIEW_ALL not in user_scopes and Scope.ADMIN_ALL not in user_scopes:
        if shipment.order.manager_id != user.id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    # Оновлюємо тільки якщо це InPost відправка
    if shipment.method not in [ShipmentMethod.INPOST_LOCKER, ShipmentMethod.INPOST_COURIER]:
        raise HTTPException(status_code=400, detail="Tracking available only for InPost shipments")
    
    if not shipment.inpost_shipment_id and not shipment.tracking_number:
        raise HTTPException(status_code=400, detail="Shipment has no InPost ID or tracking number")
    
    try:
        from modules.postal_services.service import InPostService
        
        inpost_service = InPostService(db)
        
        # Отримуємо статус з InPost API
        if shipment.inpost_shipment_id:
            # TODO: Реалізувати отримання статусу по shipment_id
            # inpost_status = await inpost_service.get_shipment_status(shipment.inpost_shipment_id)
            pass
        elif shipment.tracking_number:
            # TODO: Реалізувати отримання статусу по tracking_number
            # inpost_status = await inpost_service.get_tracking_status(shipment.tracking_number)
            pass
        
        # Оновлюємо статус на основі даних з InPost
        # shipment.status = map_inpost_status_to_shipment_status(inpost_status)
        # shipment.delivered_at = inpost_status.delivered_at if inpost_status.delivered_at else None
        
        db.commit()
        db.refresh(shipment)
        
        shipment_read = ShipmentRead.model_validate(shipment)
        shipment_read.order_number = shipment.order.order_number if shipment.order else None
        return shipment_read
        
    except Exception as e:
        logger.error(f"Failed to track shipment: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to track shipment: {str(e)}")
