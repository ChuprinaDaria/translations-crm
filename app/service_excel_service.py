"""
Сервіс для генерації Excel-файлів для відділу сервісу.

Генерує файли з:
- Обладнанням та текстилем
- Сервіруванням
- Персоналом
- Транспортними витратами
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple, Iterable

from sqlalchemy.orm import Session, joinedload

import models

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise RuntimeError("openpyxl is required for service Excel export") from e


def _load_kps_with_details(db: Session, kp_ids: List[int]) -> List[models.KP]:
    """
    Завантажує КП разом із позиціями, форматами подій та клієнтом.
    """
    return (
        db.query(models.KP)
        .options(
            joinedload(models.KP.items).joinedload(models.KPItem.item),
            joinedload(models.KP.event_formats).joinedload(models.KPEventFormat.items),
            joinedload(models.KP.client),
        )
        .filter(models.KP.id.in_(kp_ids))
        .all()
    )


def _safe_sheet_title(base: str) -> str:
    """
    Робить безпечну назву для аркуша Excel (макс. 31 символ, без заборонених символів).
    """
    invalid_chars = set(r"[]:*?/\\")
    cleaned = "".join(c for c in base if c not in invalid_chars)
    if not cleaned:
        cleaned = "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def _apply_header_style(ws, row: int, col_start: int, col_end: int):
    """Застосовує стиль заголовка до діапазону клітинок."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_cell_border(ws, row: int, col_start: int, col_end: int):
    """Застосовує рамку до діапазону клітинок."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border


def generate_service_excel(db: Session, kp_ids: List[int]) -> Tuple[bytes, str]:
    """
    Генерує Excel-файл для відділу сервісу на основі вибраних КП.

    Повертає (bytes, filename).
    """
    if not kp_ids:
        raise ValueError("Список KP ID порожній")

    kps = _load_kps_with_details(db, kp_ids)
    if not kps:
        raise ValueError("Не знайдено жодного КП для вказаних ID")

    wb = Workbook()

    # ---------- Лист "Загальна інформація" ----------
    ws_info = wb.active
    ws_info.title = "Загальна інформація"
    
    # Заголовки
    headers = ["№", "Назва КП", "Клієнт", "Дата події", "Час", "Локація", "К-сть гостей", "Формат", "Статус"]
    ws_info.append(headers)
    _apply_header_style(ws_info, 1, 1, len(headers))
    
    # Дані
    for idx, kp in enumerate(kps, 1):
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        event_date_str = ""
        if kp.event_date:
            if isinstance(kp.event_date, datetime):
                event_date_str = kp.event_date.strftime("%d.%m.%Y")
            else:
                event_date_str = str(kp.event_date)
        
        row_data = [
            idx,
            kp.title or "",
            client_name,
            event_date_str,
            kp.event_time or "",
            kp.event_location or "",
            kp.people_count or 0,
            kp.event_format or "",
            kp.status or "",
        ]
        ws_info.append(row_data)
        _apply_cell_border(ws_info, idx + 1, 1, len(headers))
    
    # Налаштування ширини колонок
    ws_info.column_dimensions['A'].width = 5
    ws_info.column_dimensions['B'].width = 35
    ws_info.column_dimensions['C'].width = 25
    ws_info.column_dimensions['D'].width = 12
    ws_info.column_dimensions['E'].width = 10
    ws_info.column_dimensions['F'].width = 35
    ws_info.column_dimensions['G'].width = 12
    ws_info.column_dimensions['H'].width = 20
    ws_info.column_dimensions['I'].width = 12

    # ---------- Лист "Обладнання" (агрегація по всіх КП) ----------
    ws_equipment = wb.create_sheet("Обладнання")
    
    # Збираємо обладнання з event_formats
    equipment_totals: Dict[str, Dict] = defaultdict(lambda: {"quantity": 0, "unit": None})
    
    for kp in kps:
        for event_format in kp.event_formats:
            for item in event_format.items:
                # Визначаємо назву
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                
                # Перевіряємо, чи це обладнання (type = 'equipment')
                item_type = item.type or (item.item.type if item.item else None)
                if item_type == 'equipment':
                    equipment_totals[name]["quantity"] += qty
                    if not equipment_totals[name]["unit"]:
                        equipment_totals[name]["unit"] = item.unit or (item.item.unit if item.item else "шт")
    
    # Також додаємо з items (якщо там є обладнання)
    for kp in kps:
        for item in kp.items:
            item_type = item.type or (item.item.type if item.item else None)
            if item_type == 'equipment':
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                equipment_totals[name]["quantity"] += qty
                if not equipment_totals[name]["unit"]:
                    equipment_totals[name]["unit"] = item.unit or (item.item.unit if item.item else "шт")
    
    # Записуємо обладнання
    eq_headers = ["№", "Назва", "Кількість", "Одиниця"]
    ws_equipment.append(eq_headers)
    _apply_header_style(ws_equipment, 1, 1, len(eq_headers))
    
    for idx, (name, data) in enumerate(sorted(equipment_totals.items()), 1):
        ws_equipment.append([
            idx,
            name,
            data["quantity"],
            data["unit"] or "шт"
        ])
        _apply_cell_border(ws_equipment, idx + 1, 1, len(eq_headers))
    
    ws_equipment.column_dimensions['A'].width = 5
    ws_equipment.column_dimensions['B'].width = 50
    ws_equipment.column_dimensions['C'].width = 12
    ws_equipment.column_dimensions['D'].width = 12

    # ---------- Лист "Сервіс" (агрегація по всіх КП) ----------
    ws_service = wb.create_sheet("Сервіс")
    
    service_totals: Dict[str, Dict] = defaultdict(lambda: {"quantity": 0, "unit": None, "price": 0})
    
    for kp in kps:
        for event_format in kp.event_formats:
            for item in event_format.items:
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                price = float(item.price or 0) if item.price else (float(item.item.price) if item.item and item.item.price else 0)
                
                item_type = item.type or (item.item.type if item.item else None)
                if item_type == 'service':
                    service_totals[name]["quantity"] += qty
                    service_totals[name]["price"] += price * qty
                    if not service_totals[name]["unit"]:
                        service_totals[name]["unit"] = item.unit or (item.item.unit if item.item else "послуга")
        
        for item in kp.items:
            item_type = item.type or (item.item.type if item.item else None)
            if item_type == 'service':
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                price = float(item.price or 0) if item.price else (float(item.item.price) if item.item and item.item.price else 0)
                service_totals[name]["quantity"] += qty
                service_totals[name]["price"] += price * qty
                if not service_totals[name]["unit"]:
                    service_totals[name]["unit"] = item.unit or (item.item.unit if item.item else "послуга")
    
    svc_headers = ["№", "Назва послуги", "Кількість", "Одиниця", "Сума, грн"]
    ws_service.append(svc_headers)
    _apply_header_style(ws_service, 1, 1, len(svc_headers))
    
    for idx, (name, data) in enumerate(sorted(service_totals.items()), 1):
        ws_service.append([
            idx,
            name,
            data["quantity"],
            data["unit"] or "послуга",
            data["price"]
        ])
        _apply_cell_border(ws_service, idx + 1, 1, len(svc_headers))
    
    ws_service.column_dimensions['A'].width = 5
    ws_service.column_dimensions['B'].width = 50
    ws_service.column_dimensions['C'].width = 12
    ws_service.column_dimensions['D'].width = 12
    ws_service.column_dimensions['E'].width = 15

    # ---------- Лист "Транспорт" ----------
    ws_transport = wb.create_sheet("Транспорт")
    
    transport_headers = ["№", "Назва КП", "Клієнт", "Локація", "Обладнання, грн", "Персонал, грн", "Всього, грн"]
    ws_transport.append(transport_headers)
    _apply_header_style(ws_transport, 1, 1, len(transport_headers))
    
    total_transport_eq = 0
    total_transport_pers = 0
    total_transport = 0
    
    for idx, kp in enumerate(kps, 1):
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        eq_cost = float(kp.transport_equipment_total or 0)
        pers_cost = float(kp.transport_personnel_total or 0)
        total_cost = float(kp.transport_total or 0) or (eq_cost + pers_cost)
        
        total_transport_eq += eq_cost
        total_transport_pers += pers_cost
        total_transport += total_cost
        
        ws_transport.append([
            idx,
            kp.title or "",
            client_name,
            kp.event_location or "",
            eq_cost,
            pers_cost,
            total_cost
        ])
        _apply_cell_border(ws_transport, idx + 1, 1, len(transport_headers))
    
    # Підсумок
    summary_row = len(kps) + 2
    ws_transport.cell(row=summary_row, column=4, value="РАЗОМ:").font = Font(bold=True)
    ws_transport.cell(row=summary_row, column=5, value=total_transport_eq).font = Font(bold=True)
    ws_transport.cell(row=summary_row, column=6, value=total_transport_pers).font = Font(bold=True)
    ws_transport.cell(row=summary_row, column=7, value=total_transport).font = Font(bold=True)
    
    ws_transport.column_dimensions['A'].width = 5
    ws_transport.column_dimensions['B'].width = 35
    ws_transport.column_dimensions['C'].width = 25
    ws_transport.column_dimensions['D'].width = 35
    ws_transport.column_dimensions['E'].width = 15
    ws_transport.column_dimensions['F'].width = 15
    ws_transport.column_dimensions['G'].width = 15

    # ---------- Окремі листи для кожного КП ----------
    for kp in kps:
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        sheet_title = client_name or kp.title or f"KP {kp.id}"
        ws_kp = wb.create_sheet(_safe_sheet_title(sheet_title))
        
        # Шапка з інформацією про захід
        ws_kp["A1"] = "Дата проведення:"
        if kp.event_date:
            if isinstance(kp.event_date, datetime):
                ws_kp["B1"] = kp.event_date.strftime("%d.%m.%Y")
            else:
                ws_kp["B1"] = str(kp.event_date)
        ws_kp["A1"].font = Font(bold=True)
        
        ws_kp["A2"] = "Час:"
        ws_kp["B2"] = kp.event_time or ""
        ws_kp["A2"].font = Font(bold=True)
        
        ws_kp["A3"] = "Замовник:"
        ws_kp["B3"] = client_name
        ws_kp["A3"].font = Font(bold=True)
        
        ws_kp["A4"] = "Місце проведення:"
        ws_kp["B4"] = kp.event_location or ""
        ws_kp["A4"].font = Font(bold=True)
        
        ws_kp["A5"] = "Формат заходу:"
        ws_kp["B5"] = kp.event_format or ""
        ws_kp["A5"].font = Font(bold=True)
        
        ws_kp["A6"] = "Кількість осіб:"
        ws_kp["B6"] = kp.people_count or 0
        ws_kp["A6"].font = Font(bold=True)
        
        current_row = 8
        
        # Обладнання для цього КП
        ws_kp.cell(row=current_row, column=1, value="ОБЛАДНАННЯ ТА СЕРВІС").font = Font(bold=True, size=12)
        current_row += 1
        
        kp_eq_headers = ["№", "Назва", "Кількість", "Одиниця"]
        for col, header in enumerate(kp_eq_headers, 1):
            ws_kp.cell(row=current_row, column=col, value=header)
        _apply_header_style(ws_kp, current_row, 1, len(kp_eq_headers))
        current_row += 1
        
        eq_idx = 1
        for event_format in kp.event_formats:
            for item in event_format.items:
                item_type = item.type or (item.item.type if item.item else None)
                if item_type == 'equipment':
                    name = item.name or (item.item.name if item.item else "Без назви")
                    qty = item.quantity or 0
                    unit = item.unit or (item.item.unit if item.item else "шт")
                    ws_kp.append([eq_idx, name, qty, unit])
                    _apply_cell_border(ws_kp, current_row, 1, 4)
                    current_row += 1
                    eq_idx += 1
        
        for item in kp.items:
            item_type = item.type or (item.item.type if item.item else None)
            if item_type == 'equipment':
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                unit = item.unit or (item.item.unit if item.item else "шт")
                ws_kp.append([eq_idx, name, qty, unit])
                _apply_cell_border(ws_kp, current_row, 1, 4)
                current_row += 1
                eq_idx += 1
        
        current_row += 1
        
        # Послуги для цього КП
        ws_kp.cell(row=current_row, column=1, value="ПОСЛУГИ").font = Font(bold=True, size=12)
        current_row += 1
        
        kp_svc_headers = ["№", "Назва", "Кількість", "Одиниця", "Сума, грн"]
        for col, header in enumerate(kp_svc_headers, 1):
            ws_kp.cell(row=current_row, column=col, value=header)
        _apply_header_style(ws_kp, current_row, 1, len(kp_svc_headers))
        current_row += 1
        
        svc_idx = 1
        for event_format in kp.event_formats:
            for item in event_format.items:
                item_type = item.type or (item.item.type if item.item else None)
                if item_type == 'service':
                    name = item.name or (item.item.name if item.item else "Без назви")
                    qty = item.quantity or 0
                    unit = item.unit or (item.item.unit if item.item else "послуга")
                    price = float(item.price or 0) if item.price else (float(item.item.price) if item.item and item.item.price else 0)
                    total = price * qty
                    ws_kp.append([svc_idx, name, qty, unit, total])
                    _apply_cell_border(ws_kp, current_row, 1, 5)
                    current_row += 1
                    svc_idx += 1
        
        for item in kp.items:
            item_type = item.type or (item.item.type if item.item else None)
            if item_type == 'service':
                name = item.name or (item.item.name if item.item else "Без назви")
                qty = item.quantity or 0
                unit = item.unit or (item.item.unit if item.item else "послуга")
                price = float(item.price or 0) if item.price else (float(item.item.price) if item.item and item.item.price else 0)
                total = price * qty
                ws_kp.append([svc_idx, name, qty, unit, total])
                _apply_cell_border(ws_kp, current_row, 1, 5)
                current_row += 1
                svc_idx += 1
        
        # Налаштування ширини колонок
        ws_kp.column_dimensions['A'].width = 20
        ws_kp.column_dimensions['B'].width = 50
        ws_kp.column_dimensions['C'].width = 12
        ws_kp.column_dimensions['D'].width = 12
        ws_kp.column_dimensions['E'].width = 15

    # Збереження у bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Ім'я файлу
    unique_dates = {
        (kp.event_date.date() if isinstance(kp.event_date, datetime) else kp.event_date)
        for kp in kps
        if kp.event_date
    }
    if len(unique_dates) == 1:
        only_date = next(iter(unique_dates))
        filename = f"service_{only_date}.xlsx"
    else:
        filename = "service.xlsx"

    return buffer.getvalue(), filename

