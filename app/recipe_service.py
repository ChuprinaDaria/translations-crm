"""
Сервіс для роботи з техкартами (рецептами) та імпорту файлу калькуляцій.

Підтримує два типи техкарт:
- catering: проста структура (страва → інгредієнти)
  Формула: вага_інгредієнта × кількість_порцій
  
- box: трирівнева структура (страва → компоненти → інгредієнти)
  Формула: вага_інгредієнта × кількість_компонентів × кількість_порцій
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, List, Tuple, Literal, Optional
from collections import defaultdict

from sqlalchemy.orm import Session, joinedload

import models
import re
from difflib import SequenceMatcher

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise RuntimeError("openpyxl is required for recipe import") from e


# ============ Константи маркерів ============
MARKER_POINT = "point"  # Маркер початку страви
MARKER_IN = "in"        # Маркер компонента боксу

# Список помилок Excel, які потрібно ігнорувати
EXCEL_ERRORS = {'#REF!', '#N/A', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!', '#ERROR!'}

# Максимальна довжина назви продукту в базі
MAX_PRODUCT_NAME_LENGTH = 500

# Спеціальна одиниця для "рядків-заголовків" всередині рецепту (підсекції/підстрави).
# Такі рядки зберігаємо як інгредієнти з вагою 0, але:
# - UI відображає їх як підзаголовки
# - розрахунок закупки їх ігнорує
GROUP_UNIT = "__group__"


def _normalize_name_for_match(name: str) -> str:
    """
    Нормалізує назву для fuzzy-match:
    - lower
    - прибирає лапки та пунктуацію
    - прибирає текст у дужках
    - стискає пробіли
    """
    if not name:
        return ""
    s = str(name).strip().lower()
    # прибираємо все в дужках
    s = re.sub(r"\([^)]*\)", " ", s)
    # уніфікуємо лапки
    s = s.replace("“", "\"").replace("”", "\"").replace("«", "\"").replace("»", "\"").replace("’", "'")
    # прибираємо пунктуацію, лишаємо букви/цифри/пробіли
    s = re.sub(r"[^0-9a-zа-яіїєґ'\\s-]+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[-_]+", " ", s)
    s = re.sub(r"\\s+", " ", s).strip()
    return s


def _tokenize(s: str) -> set[str]:
    if not s:
        return set()
    toks = {t for t in s.split() if len(t) >= 3}
    return toks


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def auto_link_recipes_to_items(
    db: Session,
    *,
    recipe_type: Optional[Literal["catering", "box"]] = None,
    threshold: float = 0.80,
    update_item_weight: bool = True,
    force_relink: bool = False,
    create_missing_items: bool = False,
    created_items_active: bool = True,
) -> Dict:
    """
    Автоматично лінкує техкарти (`recipes`) з позиціями меню (`items`) по схожості назв.

    - Записує зв'язок через `Recipe.item_id`
    - Опційно підтягуює `Item.weight`, якщо воно порожнє, з `Recipe.weight_per_portion`
    """
    result = {
        "linked": 0,
        "created_items": 0,
        "updated_item_weights": 0,
        "skipped": 0,
        "errors": [],
    }

    def _get_or_create_category(name: str) -> models.Category:
        existing = db.query(models.Category).filter(models.Category.name.ilike(name)).first()
        if existing:
            return existing
        cat = models.Category(name=name)
        db.add(cat)
        db.flush()
        return cat

    def _get_or_create_subcategory(category_id: int, name: str) -> models.Subcategory:
        existing = (
            db.query(models.Subcategory)
            .filter(models.Subcategory.category_id == category_id)
            .filter(models.Subcategory.name.ilike(name))
            .first()
        )
        if existing:
            return existing
        sub = models.Subcategory(name=name, category_id=category_id)
        db.add(sub)
        db.flush()
        return sub

    # Беремо всі items з назвою
    items = db.query(models.Item).filter(models.Item.name.isnot(None)).all()

    # Беремо всі recipes з назвою
    rq = db.query(models.Recipe).filter(models.Recipe.name.isnot(None))
    if recipe_type in ("catering", "box"):
        rq = rq.filter(models.Recipe.recipe_type == recipe_type)
    if not force_relink:
        rq = rq.filter(models.Recipe.item_id.is_(None))
    recipes = rq.all()

    # Щоб не робити кілька рецептів на один item
    used_item_ids = set(
        i for (i,) in db.query(models.Recipe.item_id).filter(models.Recipe.item_id.isnot(None)).all()
    )

    # Індексуємо recipes по токенах
    recipe_data = []
    token_index: Dict[str, List[int]] = defaultdict(list)
    for idx, r in enumerate(recipes):
        norm = _normalize_name_for_match(r.name)
        toks = _tokenize(norm)
        recipe_data.append((r, norm, toks))
        for t in toks:
            token_index[t].append(idx)

    for it in items:
        try:
            if not it.name:
                result["skipped"] += 1
                continue
            if it.id in used_item_ids and not force_relink:
                result["skipped"] += 1
                continue

            norm_it = _normalize_name_for_match(it.name)
            toks_it = _tokenize(norm_it)
            if not toks_it:
                result["skipped"] += 1
                continue

            # Кандидати: recipes, що мають спільні токени
            cand_idx = set()
            for t in toks_it:
                for ridx in token_index.get(t, []):
                    cand_idx.add(ridx)
            if not cand_idx:
                result["skipped"] += 1
                continue

            best = None
            best_score = 0.0
            for ridx in cand_idx:
                r, norm_r, toks_r = recipe_data[ridx]
                # token-jaccard як додатковий сигнал
                inter = len(toks_it & toks_r)
                union = len(toks_it | toks_r) or 1
                jacc = inter / union
                score = max(_similarity(norm_it, norm_r), jacc)
                if score > best_score:
                    best_score = score
                    best = r

            if not best or best_score < threshold:
                result["skipped"] += 1
                continue

            # Лінкуємо
            best.item_id = it.id
            used_item_ids.add(it.id)
            result["linked"] += 1

            # Підтягуємо weight в item (якщо пусто)
            if update_item_weight and (it.weight is None or str(it.weight).strip() == "") and best.weight_per_portion:
                w = float(best.weight_per_portion)
                it.weight = str(int(w)) if abs(w - int(w)) < 1e-6 else str(w)
                result["updated_item_weights"] += 1
        except Exception as e:
            result["errors"].append(f"Помилка лінку item '{getattr(it,'name',None)}': {e}")

    # Додаткова перевірка: якщо в меню немає страви, але в техкартах є — перевіряємо існуючі страви за 80% схожістю назв
    # НЕ створюємо нові страви, тільки перевіряємо існуючі та підв'язуємо до них
    # Використовуємо threshold 0.80 для виявлення можливих збігів
    similarity_threshold = 0.80
    try:
        # existing items для перевірки схожості
        all_items = db.query(models.Item).filter(models.Item.name.isnot(None)).all()
        
        # беремо рецепти, які ще не зв'язані (або всі, якщо force_relink)
        rq2 = db.query(models.Recipe).filter(models.Recipe.name.isnot(None))
        if recipe_type in ("catering", "box"):
            rq2 = rq2.filter(models.Recipe.recipe_type == recipe_type)
        if not force_relink:
            rq2 = rq2.filter(models.Recipe.item_id.is_(None))

        for r in rq2.all():
            norm_r = _normalize_name_for_match(r.name)
            if not norm_r:
                continue
            
            # Шукаємо найкращий збіг за схожістю (мінімум 80%)
            best_item = None
            best_score = 0.0
            
            for item in all_items:
                if item.id in used_item_ids and not force_relink:
                    continue
                    
                norm_item = _normalize_name_for_match(item.name)
                if not norm_item:
                    continue
                
                # Обчислюємо схожість
                similarity = _similarity(norm_r, norm_item)
                
                if similarity >= similarity_threshold and similarity > best_score:
                    best_score = similarity
                    best_item = item
            
            if best_item and best_item.id not in used_item_ids:
                # Зв'язуємо техкарту з найкращим збігом
                r.item_id = best_item.id
                used_item_ids.add(best_item.id)
                result["linked"] += 1
                # Зберігаємо інформацію про рівень схожості в notes (якщо потрібно для UI)
                # Можна додати окреме поле link_similarity_score в майбутньому
            else:
                result["skipped"] += 1
    except Exception as e:
        result["errors"].append(f"Additional similarity match linking failed: {e}")

    db.flush()
    return result


def is_cooking_step(text: str) -> bool:
    """
    Перевіряє, чи є текст кроком приготування, а не назвою інгредієнта.
    
    Кроки приготування зазвичай:
    - Починаються з цифри та крапки (1., 2., 3.)
    - Містять дієслова у наказовому способі (Змішайте, Додайте, Подавайте)
    - Довші за 100 символів
    """
    if not text or not isinstance(text, str):
        return False
    
    text = text.strip()
    
    # Перевірка на крок з номером (1., 2., 3. і т.д.)
    import re
    if re.match(r'^\d+\.\s', text):
        return True
    
    # Перевірка на довгий текст з дієсловами приготування
    cooking_verbs = [
        'змішайте', 'додайте', 'подавайте', 'запікайте', 'варіть', 
        'смажте', 'нарізати', 'нарізайте', 'залиште', 'охолодіть',
        'перемішайте', 'помістіть', 'сформуйте', 'обверніть', 'викладіть',
        'накрийте', 'зачекайте', 'дайте', 'поставте'
    ]
    
    text_lower = text.lower()
    if len(text) > 100 and any(verb in text_lower for verb in cooking_verbs):
        return True
    
    return False


def clean_product_name(name: str) -> str:
    """
    Очищає та обрізає назву продукту для збереження в базі.
    """
    if not name:
        return ""
    
    # Видаляємо зайві пробіли
    cleaned = str(name).strip()
    
    # Обрізаємо до максимальної довжини
    if len(cleaned) > MAX_PRODUCT_NAME_LENGTH:
        cleaned = cleaned[:MAX_PRODUCT_NAME_LENGTH - 3] + "..."
    
    return cleaned


def safe_float(value, default: float = 0.0) -> Optional[float]:
    """
    Безпечно конвертує значення в float, ігноруючи помилки Excel.
    
    Args:
        value: Значення з клітинки Excel
        default: Значення за замовчуванням якщо конвертація неможлива
    
    Returns:
        float або None
    """
    if value is None:
        return None
    
    # Якщо це вже число - повертаємо
    if isinstance(value, (int, float)):
        return float(value)
    
    # Конвертуємо в строку для перевірки
    str_value = str(value).strip()
    
    # Перевіряємо на помилки Excel
    if str_value.upper() in EXCEL_ERRORS or str_value.startswith('#'):
        return default
    
    # Якщо порожня строка
    if not str_value:
        return None
    
    # Спробуємо конвертувати
    try:
        return float(str_value.replace(',', '.'))
    except (ValueError, TypeError):
        return default


def parse_portion_weight(value) -> Optional[float]:
    """
    Парсить "вагу порції" з колонки B в Excel.

    У файлі "Оновлена закупка 2024" вага порції часто вказана як рядок типу:
    - "120/120" (сума частин)
    - "200/25"
    - "150"
    - інколи діапазон "120-130"
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    # Дістаємо всі числа (з десятковими, комами)
    nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", s)]
    if not nums:
        return None

    # Якщо кілька частин порції — сумуємо
    if "/" in s or "+" in s:
        return float(sum(nums))

    # Якщо діапазон — беремо середнє
    if "-" in s and len(nums) == 2:
        return float((nums[0] + nums[1]) / 2.0)

    # Інакше — перше число
    return float(nums[0])


# ============ Імпорт техкарт ============

def import_calculations_file(
    db: Session, 
    file_content: bytes, 
    recipe_type: Literal["catering", "box"] = "catering"
) -> Dict:
    """
    Імпортує файл калькуляцій (техкарт) у форматі Excel.
    
    Структура файлу:
    - Лист "калькуляции": техкарти страв
      - Колонка A: категорія або кількість
      - Колонка C: назва страви/інгредієнта
      - Колонка E: вага в грамах
      - Колонка G: маркер 'point' для початку техкарти, 'in' для компонента (бокси)
    
    - Лист "список продуктов": словник продуктів
    
    Args:
        db: Сесія бази даних
        file_content: Вміст Excel файлу
        recipe_type: Тип техкарт ('catering' або 'box')
    
    Повертає словник з результатами імпорту.
    """
    wb = load_workbook(BytesIO(file_content), data_only=True)
    
    result = {
        "recipes_imported": 0,
        "components_imported": 0,
        "products_imported": 0,
        "errors": [],
        "recipe_type": recipe_type
    }
    
    # Імпорт техкарт
    # У різних файлах назва листа може бути "калькуляции"/"калькуляції"/"калькуляціі" тощо.
    sheet = None
    preferred_names = {"калькуляции", "калькуляції", "калькуляціі", "калькуляции "}
    for name in wb.sheetnames:
        name_norm = str(name).strip().lower()
        if name_norm in preferred_names or "калькуля" in name_norm:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[0]
    
    if recipe_type == "box":
        recipes_result = _import_box_recipes_from_sheet(db, sheet)
        result["components_imported"] = recipes_result.get("components_count", 0)
    else:
        recipes_result = _import_catering_recipes_from_sheet(db, sheet)
    
    result["recipes_imported"] = recipes_result["count"]
    result["errors"].extend(recipes_result["errors"])

    # Після імпорту пробуємо автоматично зв'язати техкарти з позиціями меню (items)
    try:
        link_res = auto_link_recipes_to_items(db, recipe_type=recipe_type, create_missing_items=True)
        result["items_linked"] = link_res.get("linked", 0)
        result["items_created"] = link_res.get("created_items", 0)
        result["items_weight_updated"] = link_res.get("updated_item_weights", 0)
        if link_res.get("errors"):
            result["errors"].extend(link_res["errors"])
    except Exception as e:
        result["errors"].append(f"Auto-link items skipped/failed: {e}")
    
    # Імпорт продуктів
    if "список продуктов" in wb.sheetnames:
        products_sheet = wb["список продуктов"]
        products_result = _import_products_from_sheet(db, products_sheet)
        result["products_imported"] = products_result["count"]
        result["errors"].extend(products_result["errors"])
    
    db.commit()
    return result


def _normalize_marker(value) -> Optional[str]:
    """
    Нормалізує маркер з Excel (колонка G):
    - прибирає пробіли
    - переводить у lower-case
    Повертає None, якщо значення порожнє або не є строкою.
    """
    if value is None or not isinstance(value, str):
        return None
    s = value.strip().lower()
    return s or None


def _sheet_has_markers(sheet) -> bool:
    """
    Перевіряє, чи є в листі маркери 'point'/'in' у колонці G.
    Якщо їх немає (як у файлах типу "Оновлена закупка 2024"), працюємо в markerless-режимі.
    """
    try:
        max_row = sheet.max_row
    except Exception:
        return False
    for row in range(1, max_row + 1):
        v = sheet.cell(row=row, column=7).value
        m = _normalize_marker(v)
        if m in {MARKER_POINT, MARKER_IN}:
            return True
    return False


def _is_header_row(col_b, col_c, col_d, col_e) -> bool:
    """
    Визначає заголовок таблиці.
    """
    header_tokens = {"вихід, грам", "назва", "продукти", "вага у грамах"}
    for v in (col_b, col_c, col_d, col_e):
        if isinstance(v, str) and v.strip().lower() in header_tokens:
            return True
    return False


def _merged_rows_covering_a_to_e(sheet) -> set[int]:
    """
    Повертає множину номерів рядків, де є merge-діапазон, що покриває колонки A..E.
    У файлі "Оновлена закупка 2024 (8)" такі рядки відповідають "зеленим" категоріям.
    """
    rows: set[int] = set()
    try:
        ranges = list(getattr(sheet, "merged_cells", []).ranges)
    except Exception:
        ranges = []
    for rng in ranges:
        try:
            if rng.min_row == rng.max_row and rng.min_col <= 1 and rng.max_col >= 5:
                rows.add(int(rng.min_row))
        except Exception:
            continue
    return rows


def _is_category_row(row_idx: int, merged_category_rows: set[int], col_a, col_b, col_c, col_d, col_e) -> bool:
    """
    Категорія — це "зелений" рядок (merge A:E), де фактично є тільки назва категорії,
    а B/C/D/E — порожні.
    """
    if row_idx not in merged_category_rows:
        return False
    if not (col_a and isinstance(col_a, str) and col_a.strip()):
        return False
    if any(v not in (None, "") for v in (col_b, col_c, col_d, col_e)):
        return False
    return True


def _is_point_category_row(marker: Optional[str], col_a, col_b, col_c, col_d, col_e) -> bool:
    """
    Частина файлів використовує `point` і для категорії, і для страви.
    Категорія при цьому виглядає як: A = текст, а B/C/D/E порожні.
    (Merge A:E — бажаний сигнал, але не завжди присутній.)
    """
    if marker != MARKER_POINT:
        return False
    if not (col_a and isinstance(col_a, str) and col_a.strip()):
        return False
    if any(v not in (None, "") for v in (col_b, col_c, col_d, col_e)):
        return False
    return True


def _portion_weight_from_b_or_e(col_b, col_e) -> Optional[float]:
    """
    Визначає вагу порції (в грамах) для страви.

    Правило:
    - у "Оновлена закупка 2024" колонка B — "Вихід, грам", але інколи там буває "12шт."
      (не грами). У такому випадку беремо вагу з колонки E.
    """
    parsed_b = parse_portion_weight(col_b)
    e_val = safe_float(col_e, None)

    # Якщо B — строка з ознаками "не грами" (шт/pcs/порц) — ігноруємо parsed_b.
    if isinstance(col_b, str):
        s = col_b.strip().lower()
        if any(tok in s for tok in ("шт", "pcs", "pc", "порц", "порції", "порция")):
            return e_val if e_val is not None else parsed_b

    # Додаткова евристика: якщо B дає дуже мале число, а E — велике,
    # то B, ймовірно, не вага в грамах (напр. "12шт.").
    if parsed_b is not None and e_val is not None:
        if parsed_b < 30 and e_val >= 100:
            return e_val

    return parsed_b if parsed_b is not None else e_val


def _pick_weight_cell(col_e, col_f):
    """
    Деякі файли можуть містити вагу не в колонці E, а в F.
    Беремо перше непорожнє значення.
    """
    if col_e not in (None, ""):
        return col_e
    if col_f not in (None, ""):
        return col_f
    return None


def _import_catering_recipes_from_sheet(db: Session, sheet) -> Dict:
    """
    Імпортує техкарти для кейтерінгу (проста структура).
    Страва → Інгредієнти
    """
    result = {"count": 0, "errors": []}
    
    max_row = sheet.max_row
    current_recipe = None
    current_category = None
    ingredient_index = 0
    merged_category_rows = _merged_rows_covering_a_to_e(sheet)

    for row in range(1, max_row + 1):
        col_a = sheet.cell(row=row, column=1).value
        col_b = sheet.cell(row=row, column=2).value
        col_c = sheet.cell(row=row, column=3).value
        col_d = sheet.cell(row=row, column=4).value
        col_e = sheet.cell(row=row, column=5).value
        col_f = sheet.cell(row=row, column=6).value
        col_g = sheet.cell(row=row, column=7).value
        marker = _normalize_marker(col_g)
        
        # Пропускаємо заголовки / порожні рядки
        if _is_header_row(col_b, col_c, col_d, col_e):
            continue
        if all(v in (None, "") for v in (col_a, col_b, col_c, col_d, col_e, col_f, col_g)):
            continue

        # Категорія: зелений рядок (merge A:E)
        if _is_category_row(row, merged_category_rows, col_a, col_b, col_c, col_d, col_e):
            current_category = str(col_a).strip()
            continue

        # Категорія (fallback): marker 'point' + тільки A заповнена
        if _is_point_category_row(marker, col_a, col_b, col_c, col_d, col_e):
            current_category = str(col_a).strip()
            continue
        
        # Категорія (marker-based): A + marker 'point'
        if col_a and isinstance(col_a, str) and marker == MARKER_POINT:
            current_category = str(col_a).strip()
            continue
        
        # Визначаємо "старт нової страви"
        is_new_recipe = False
        # Marker-based: marker 'point' + назва в C
        if marker == MARKER_POINT and col_c not in (None, ""):
            is_new_recipe = True
        # Markerless: на рядку страви заповнені B, C, E; D зазвичай порожній
        elif col_c not in (None, "") and col_b not in (None, "") and _pick_weight_cell(col_e, col_f) is not None and col_d in (None, ""):
            is_new_recipe = True

        if is_new_recipe and col_c:
            # Зберігаємо попередню страву
            if current_recipe:
                try:
                    db.add(current_recipe)
                    result["count"] += 1
                except Exception as e:
                    result["errors"].append(f"Помилка збереження страви '{current_recipe.name}': {str(e)}")
            
            # Перевіряємо, чи така страва вже існує
            existing = db.query(models.Recipe).filter(
                models.Recipe.name == str(col_c).strip(),
                models.Recipe.recipe_type == "catering"
            ).first()
            
            if existing:
                current_recipe = existing
                # Очищаємо старі інгредієнти
                db.query(models.RecipeIngredient).filter(
                    models.RecipeIngredient.recipe_id == existing.id
                ).delete()
                # Оновлюємо базові поля (вага/категорія) при повторному імпорті
                existing.category = current_category
                # Вихід страви завжди в колонці B (не змішуємо з вагою інгредієнтів з колонки E)
                existing.weight_per_portion = parse_portion_weight(col_b)
            else:
                current_recipe = models.Recipe(
                    name=str(col_c).strip(),
                    category=current_category,
                    # Вихід страви завжди в колонці B (напр. "120/120").
                    # Колонка E - це вага інгредієнтів, не вихід страви.
                    weight_per_portion=parse_portion_weight(col_b),
                    recipe_type="catering"
                )
            ingredient_index = 0
            continue

        # Підсекція/підстрава всередині страви (напр. "Ростбіф" / "Маринад:").
        # Це рядок з текстом у C, де D/E порожні; нижче йдуть інгредієнти у D/E.
        weight_cell = _pick_weight_cell(col_e, col_f)
        if (
            current_recipe
            and marker not in {MARKER_POINT, MARKER_IN}
            and col_d in (None, "")
            and isinstance(col_c, str)
            and col_c.strip()
            and weight_cell is None
        ):
            label = col_c.strip().rstrip(":").strip()
            if label and not is_cooking_step(label):
                # Вихід підсекції завжди в колонці B (не змішуємо з вагою інгредієнтів з колонки E)
                group_weight = parse_portion_weight(col_b) or 0.0
                group_row = models.RecipeIngredient(
                    product_name=clean_product_name(label),
                    weight_per_portion=group_weight,
                    unit=GROUP_UNIT,
                    order_index=ingredient_index,
                )
                current_recipe.ingredients.append(group_row)
                ingredient_index += 1
            continue

        # Якщо немає маркера і є назва - це інгредієнт
        if current_recipe and weight_cell is not None:
            # У цьому Excel назва інгредієнта зазвичай в колонці D,
            # а колонка C може містити підзаголовки типу "Маринад:".
            raw_name = col_d if col_d not in [None, ""] else col_c
            if raw_name in [None, ""]:
                continue

            product_name = str(raw_name).strip()

            # Підзаголовки груп інгредієнтів (напр. "Маринад:") — зберігаємо як підсекцію
            if raw_name == col_c and product_name.endswith(":") and col_d in [None, ""]:
                label = product_name.rstrip(":").strip()
                if label and not is_cooking_step(label):
                    # Вихід підсекції завжди в колонці B
                    group_weight = parse_portion_weight(col_b) or 0.0
                    group_row = models.RecipeIngredient(
                        product_name=clean_product_name(label),
                        weight_per_portion=group_weight,
                        unit=GROUP_UNIT,
                        order_index=ingredient_index,
                    )
                    current_recipe.ingredients.append(group_row)
                    ingredient_index += 1
                continue
            
            # Пропускаємо кроки приготування
            if is_cooking_step(product_name):
                continue
            
            weight = safe_float(weight_cell, None)
            if weight is not None:
                # Рядки-підзаголовки без двокрапки (напр. "Ростбіф") інколи мають 0 у вазі в колонці E.
                # У такому випадку зберігаємо як підсекцію, але вихід беремо з колонки B.
                if (col_d in [None, ""]) and (raw_name == col_c) and float(weight) == 0.0:
                    label = product_name.rstrip(":").strip()
                    if label and not is_cooking_step(label):
                        # Вихід підсекції завжди в колонці B
                        group_weight = parse_portion_weight(col_b) or 0.0
                        group_row = models.RecipeIngredient(
                            product_name=clean_product_name(label),
                            weight_per_portion=group_weight,
                            unit=GROUP_UNIT,
                            order_index=ingredient_index,
                        )
                        current_recipe.ingredients.append(group_row)
                        ingredient_index += 1
                    continue
                ingredient = models.RecipeIngredient(
                    product_name=clean_product_name(product_name),
                    weight_per_portion=weight,
                    unit="г",
                    order_index=ingredient_index
                )
                current_recipe.ingredients.append(ingredient)
                ingredient_index += 1
    
    # Зберігаємо останню страву
    if current_recipe:
        try:
            db.add(current_recipe)
            result["count"] += 1
        except Exception as e:
            result["errors"].append(f"Помилка збереження страви '{current_recipe.name}': {str(e)}")
    
    return result


def _import_box_recipes_from_sheet(db: Session, sheet) -> Dict:
    """
    Імпортує техкарти для боксів (трирівнева структура).
    Страва → Компоненти → Інгредієнти
    
    Маркери:
    - 'point' - початок страви
    - 'in' - початок компонента боксу
    """
    result = {"count": 0, "components_count": 0, "errors": []}
    
    max_row = sheet.max_row
    current_recipe = None
    current_component = None
    current_category = None
    component_index = 0
    ingredient_index = 0
    merged_category_rows = _merged_rows_covering_a_to_e(sheet)
    
    for row in range(1, max_row + 1):
        col_a = sheet.cell(row=row, column=1).value
        col_b = sheet.cell(row=row, column=2).value
        col_c = sheet.cell(row=row, column=3).value
        col_d = sheet.cell(row=row, column=4).value
        col_e = sheet.cell(row=row, column=5).value
        col_f = sheet.cell(row=row, column=6).value
        col_g = sheet.cell(row=row, column=7).value
        marker = _normalize_marker(col_g)
        
        # Пропускаємо заголовки / порожні рядки
        if _is_header_row(col_b, col_c, col_d, col_e):
            continue
        if all(v in (None, "") for v in (col_a, col_b, col_c, col_d, col_e, col_f, col_g)):
            continue
        
        # Категорія: зелений рядок (merge A:E)
        if _is_category_row(row, merged_category_rows, col_a, col_b, col_c, col_d, col_e):
            current_category = str(col_a).strip()
            continue

        # Категорія (fallback): marker 'point' + тільки A заповнена
        if _is_point_category_row(marker, col_a, col_b, col_c, col_d, col_e):
            current_category = str(col_a).strip()
            continue

        # Категорія (marker-based): A + marker 'point'
        if col_a and isinstance(col_a, str) and marker == MARKER_POINT:
            current_category = str(col_a).strip()
            continue
        
        # Якщо є маркер 'point' і назва - це нова страва (бокс)
        if marker == MARKER_POINT and col_c:
            # Зберігаємо попередню страву
            if current_recipe:
                try:
                    db.add(current_recipe)
                    result["count"] += 1
                except Exception as e:
                    result["errors"].append(f"Помилка збереження боксу '{current_recipe.name}': {str(e)}")
            
            # Перевіряємо, чи такий бокс вже існує
            existing = db.query(models.Recipe).filter(
                models.Recipe.name == str(col_c).strip(),
                models.Recipe.recipe_type == "box"
            ).first()
            
            if existing:
                current_recipe = existing
                # Очищаємо старі компоненти
                db.query(models.RecipeComponent).filter(
                    models.RecipeComponent.recipe_id == existing.id
                ).delete()
                # Також чистимо прямі інгредієнти (якщо були)
                db.query(models.RecipeIngredient).filter(
                    models.RecipeIngredient.recipe_id == existing.id
                ).delete()
                # Оновлюємо базові поля при повторному імпорті
                existing.category = current_category
                # Вихід страви завжди в колонці B (не змішуємо з вагою інгредієнтів з колонки E)
                existing.weight_per_portion = parse_portion_weight(col_b)
            else:
                current_recipe = models.Recipe(
                    name=str(col_c).strip(),
                    category=current_category,
                    # Вихід страви завжди в колонці B (не змішуємо з вагою інгредієнтів з колонки E)
                    weight_per_portion=parse_portion_weight(col_b),
                    recipe_type="box"
                )
            
            current_component = None
            component_index = 0
            continue
        
        # Якщо є маркер 'in' - це компонент боксу
        if marker == MARKER_IN and col_c and current_recipe:
            current_component = models.RecipeComponent(
                name=str(col_c).strip(),
                quantity_per_portion=safe_float(col_a, 1.0) or 1.0,
                # Вихід компонента завжди в колонці B (не змішуємо з вагою інгредієнтів з колонки E)
                weight_per_portion=parse_portion_weight(col_b),
                order_index=component_index
            )
            current_recipe.components.append(current_component)
            result["components_count"] += 1
            component_index += 1
            ingredient_index = 0
            continue

        # Підсекція всередині компонента (або всередині боксу без компонента):
        # текст у C, D/E порожні. Нижче йдуть інгредієнти у D/E.
        weight_cell = _pick_weight_cell(col_e, col_f)
        if (
            marker not in {MARKER_POINT, MARKER_IN}
            and col_d in (None, "")
            and isinstance(col_c, str)
            and col_c.strip()
            and weight_cell is None
        ):
            label = col_c.strip().rstrip(":").strip()
            if label and not is_cooking_step(label):
                # Вихід підсекції завжди в колонці B
                group_weight = parse_portion_weight(col_b) or 0.0
                if current_component:
                    group_row = models.RecipeComponentIngredient(
                        product_name=clean_product_name(label),
                        weight_per_unit=group_weight,
                        unit=GROUP_UNIT,
                        order_index=ingredient_index,
                    )
                    current_component.ingredients.append(group_row)
                    ingredient_index += 1
                    continue
                if current_recipe and not current_component:
                    group_row = models.RecipeIngredient(
                        product_name=clean_product_name(label),
                        weight_per_portion=group_weight,
                        unit=GROUP_UNIT,
                        order_index=ingredient_index,
                    )
                    current_recipe.ingredients.append(group_row)
                    ingredient_index += 1
                    continue
        
        # Якщо є компонент і є назва/вага - це інгредієнт компонента
        if current_component and weight_cell is not None:
            raw_name = col_d if col_d not in [None, ""] else col_c
            if raw_name in [None, ""]:
                continue
            product_name = str(raw_name).strip()

            if raw_name == col_c and product_name.endswith(":") and col_d in [None, ""]:
                label = product_name.rstrip(":").strip()
                if label and not is_cooking_step(label):
                    # Вихід підсекції завжди в колонці B
                    group_weight = parse_portion_weight(col_b) or 0.0
                    group_row = models.RecipeComponentIngredient(
                        product_name=clean_product_name(label),
                        weight_per_unit=group_weight,
                        unit=GROUP_UNIT,
                        order_index=ingredient_index,
                    )
                    current_component.ingredients.append(group_row)
                    ingredient_index += 1
                continue
            
            # Пропускаємо кроки приготування
            if is_cooking_step(product_name):
                continue
            
            weight = safe_float(weight_cell, None)
            if weight is not None:
                if (col_d in [None, ""]) and (raw_name == col_c) and float(weight) == 0.0:
                    label = product_name.rstrip(":").strip()
                    if label and not is_cooking_step(label):
                        # Вихід підсекції завжди в колонці B (не з колонки E, де вага інгредієнта)
                        group_weight = parse_portion_weight(col_b) or 0.0
                        group_row = models.RecipeComponentIngredient(
                            product_name=clean_product_name(label),
                            weight_per_unit=group_weight,
                            unit=GROUP_UNIT,
                            order_index=ingredient_index,
                        )
                        current_component.ingredients.append(group_row)
                        ingredient_index += 1
                    continue
                ingredient = models.RecipeComponentIngredient(
                    product_name=clean_product_name(product_name),
                    weight_per_unit=weight,
                    unit="г",
                    order_index=ingredient_index
                )
                current_component.ingredients.append(ingredient)
                ingredient_index += 1
        
        # Якщо немає компонента, але є страва і інгредієнт - додаємо як прямий інгредієнт
        elif current_recipe and not current_component and weight_cell is not None:
            raw_name = col_d if col_d not in [None, ""] else col_c
            if raw_name in [None, ""]:
                continue
            product_name = str(raw_name).strip()

            if raw_name == col_c and product_name.endswith(":") and col_d in [None, ""]:
                label = product_name.rstrip(":").strip()
                if label and not is_cooking_step(label):
                    # Вихід підсекції завжди в колонці B
                    group_weight = parse_portion_weight(col_b) or 0.0
                    group_row = models.RecipeIngredient(
                        product_name=clean_product_name(label),
                        weight_per_portion=group_weight,
                        unit=GROUP_UNIT,
                        order_index=ingredient_index,
                    )
                    current_recipe.ingredients.append(group_row)
                    ingredient_index += 1
                continue
            
            # Пропускаємо кроки приготування
            if is_cooking_step(product_name):
                continue
            
            weight = safe_float(weight_cell, None)
            if weight is not None:
                if (col_d in [None, ""]) and (raw_name == col_c) and float(weight) == 0.0:
                    label = product_name.rstrip(":").strip()
                    if label and not is_cooking_step(label):
                        # Вихід підсекції завжди в колонці B (не з колонки E, де вага інгредієнта)
                        group_weight = parse_portion_weight(col_b) or 0.0
                        group_row = models.RecipeIngredient(
                            product_name=clean_product_name(label),
                            weight_per_portion=group_weight,
                            unit=GROUP_UNIT,
                            order_index=ingredient_index,
                        )
                        current_recipe.ingredients.append(group_row)
                        ingredient_index += 1
                    continue
                ingredient = models.RecipeIngredient(
                    product_name=clean_product_name(product_name),
                    weight_per_portion=weight,
                    unit="г",
                    order_index=ingredient_index
                )
                current_recipe.ingredients.append(ingredient)
                ingredient_index += 1
    
    # Зберігаємо останню страву
    if current_recipe:
        try:
            db.add(current_recipe)
            result["count"] += 1
        except Exception as e:
            result["errors"].append(f"Помилка збереження боксу '{current_recipe.name}': {str(e)}")
    
    return result


def _import_products_from_sheet(db: Session, sheet) -> Dict:
    """Імпортує продукти з листа 'список продуктов'."""
    result = {"count": 0, "errors": [], "skipped": 0}
    
    max_row = sheet.max_row
    current_category = None
    current_subcategory = None
    
    # Отримуємо всі існуючі продукти одразу (для швидкості)
    existing_products = set(
        name for (name,) in db.query(models.Product.name).all()
    )
    # Також тримаємо список нових продуктів з цього імпорту
    added_in_this_import = set()
    
    for row in range(1, max_row + 1):
        col_a = sheet.cell(row=row, column=1).value
        
        if not col_a:
            continue
        
        col_a = str(col_a).strip()
        
        # Пропускаємо заголовки
        if col_a == "Позначення":
            continue
        
        # Визначаємо категорії (великими літерами)
        if col_a.isupper() and len(col_a) > 2:
            if any(char.isalpha() for char in col_a):
                if current_category is None or col_a in ["РИБНИЙ", "М'ЯСНИЙ", "ОВОЧІ", "МОЛОЧНИЙ", "БАКАЛІЯ", "ІНШЕ"]:
                    current_category = col_a
                    current_subcategory = None
                else:
                    current_subcategory = col_a
                continue
        
        # Пропускаємо якщо продукт вже існує або вже додано в цьому імпорті
        if col_a in existing_products or col_a in added_in_this_import:
            result["skipped"] += 1
            continue
        
        # Це новий продукт
        product = models.Product(
            name=col_a,
            category=current_category,
            subcategory=current_subcategory,
            unit="г"
        )
        try:
            db.add(product)
            added_in_this_import.add(col_a)
            result["count"] += 1
        except Exception as e:
            result["errors"].append(f"Помилка збереження продукту '{col_a}': {str(e)}")
    
    return result


# ============ Отримання техкарт ============

def get_all_recipes(
    db: Session, 
    recipe_type: Optional[Literal["catering", "box"]] = None
) -> List[models.Recipe]:
    """
    Отримує всі техкарти.
    
    Args:
        db: Сесія бази даних
        recipe_type: Фільтр за типом (None = всі)
    """
    query = db.query(models.Recipe).options(
        joinedload(models.Recipe.ingredients),
        joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients),
        joinedload(models.Recipe.item)
    )
    
    if recipe_type:
        query = query.filter(models.Recipe.recipe_type == recipe_type)
    
    return query.all()


def get_recipe_by_name(
    db: Session, 
    name: str,
    recipe_type: Optional[Literal["catering", "box"]] = None
) -> Optional[models.Recipe]:
    """
    Шукає техкарту за назвою страви (нечутливий до регістру).
    
    Args:
        db: Сесія бази даних
        name: Назва страви
        recipe_type: Фільтр за типом (None = будь-який)
    """
    query = db.query(models.Recipe).options(
        joinedload(models.Recipe.ingredients),
        joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients)
    ).filter(models.Recipe.name.ilike(f"%{name}%"))
    
    if recipe_type:
        query = query.filter(models.Recipe.recipe_type == recipe_type)
    
    return query.first()


def get_recipe_by_id(db: Session, recipe_id: int) -> Optional[models.Recipe]:
    """Отримує техкарту за ID."""
    return db.query(models.Recipe).options(
        joinedload(models.Recipe.ingredients),
        joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients)
    ).filter(models.Recipe.id == recipe_id).first()


# ============ Розрахунок закупки ============

def calculate_purchase_from_kps(
    db: Session, 
    kp_ids: List[int],
    calculation_type: Optional[Literal["catering", "box", "auto"]] = "auto"
) -> Dict[str, Dict]:
    """
    Розраховує закупку продуктів на основі вибраних КП.
    
    Алгоритм залежить від типу:
    
    **Кейтерінг (catering):**
    1. Беремо страви з КП
    2. Для кожної страви шукаємо техкарту
    3. Множимо: вага_інгредієнта × кількість_порцій
    
    **Бокси (box):**
    1. Беремо страви (бокси) з КП
    2. Для кожного боксу шукаємо техкарту
    3. Множимо: вага_інгредієнта × кількість_компонентів × кількість_порцій
    
    **Auto:**
    - Визначає тип автоматично на основі event_group КП
    
    Args:
        db: Сесія бази даних
        kp_ids: Список ID КП
        calculation_type: Тип розрахунку
    
    Повертає: {product_name: {"quantity": float, "unit": str}}
    """
    # Завантажуємо КП з позиціями
    kps = (
        db.query(models.KP)
        .options(
            joinedload(models.KP.items).joinedload(models.KPItem.item),
        )
        .filter(models.KP.id.in_(kp_ids))
        .all()
    )
    
    if not kps:
        return {
            "products": {},
            "dishes_without_recipe": [],
            "total_dishes": 0,
            "dishes_with_recipe": 0,
            "calculation_type": calculation_type
        }
    
    # Визначаємо тип розрахунку автоматично
    if calculation_type == "auto":
        # Перевіряємо event_group першого КП
        first_kp = kps[0]
        if first_kp.event_group and "бокс" in first_kp.event_group.lower():
            calculation_type = "box"
        else:
            calculation_type = "catering"
    
    # Збираємо страви та їх кількості
    dish_quantities: Dict[str, float] = defaultdict(float)
    item_quantities: Dict[int, float] = defaultdict(float)
    
    for kp in kps:
        for kp_item in kp.items:
            # Визначаємо назву страви
            if kp_item.item and kp_item.item.name:
                dish_name = kp_item.item.name
            elif kp_item.name:
                dish_name = kp_item.name
            else:
                continue
            
            # Визначаємо тип - нас цікавлять тільки страви (menu)
            item_type = None
            if kp_item.type:
                item_type = kp_item.type
            elif kp_item.item and hasattr(kp_item.item, 'type'):
                item_type = kp_item.item.type
            
            # Пропускаємо обладнання та сервіс
            if item_type in ['equipment', 'service']:
                continue
            
            qty = kp_item.quantity or 0
            dish_quantities[dish_name] += qty
            if kp_item.item_id:
                item_quantities[int(kp_item.item_id)] += qty
    
    # Розраховуємо продукти
    product_totals: Dict[str, Dict] = defaultdict(lambda: {"quantity": 0.0, "unit": "г"})
    dishes_without_recipe = []
    dishes_with_recipe = 0

    # 1) Швидкий шлях: по item_id (якщо техкарта зв'язана з item)
    recipes_by_item_id: Dict[int, models.Recipe] = {}
    if item_quantities:
        q = (
            db.query(models.Recipe)
            .options(
                joinedload(models.Recipe.ingredients),
                joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients),
            )
            .filter(models.Recipe.item_id.in_(list(item_quantities.keys())))
        )
        if calculation_type in ("catering", "box"):
            q = q.filter(models.Recipe.recipe_type == calculation_type)
        for r in q.all():
            if r.item_id is not None:
                recipes_by_item_id[int(r.item_id)] = r

    for item_id, quantity in item_quantities.items():
        recipe = recipes_by_item_id.get(int(item_id))
        if recipe:
            dishes_with_recipe += 1
            if recipe.recipe_type == "box" and recipe.components:
                _calculate_box_products(recipe, quantity, product_totals)
            else:
                _calculate_catering_products(recipe, quantity, product_totals)

    # 2) Fallback: по назві (для custom items або не зв'язаних items)
    for dish_name, quantity in dish_quantities.items():
        recipe = get_recipe_by_name(db, dish_name, calculation_type)
        if not recipe:
            recipe = get_recipe_by_name(db, dish_name)
        if recipe:
            # Не подвоюємо, якщо вже порахували по item_id (в такому разі item_id-страви теж мають dish_name)
            if recipe.item_id and int(recipe.item_id) in item_quantities:
                continue
            dishes_with_recipe += 1
            if recipe.recipe_type == "box" and recipe.components:
                _calculate_box_products(recipe, quantity, product_totals)
            else:
                _calculate_catering_products(recipe, quantity, product_totals)
        else:
            dishes_without_recipe.append(dish_name)
    
    return {
        "products": dict(product_totals),
        "dishes_without_recipe": dishes_without_recipe,
        "total_dishes": len(dish_quantities),
        "dishes_with_recipe": dishes_with_recipe,
        "calculation_type": calculation_type
    }


def _calculate_catering_products(
    recipe: models.Recipe, 
    quantity: float, 
    product_totals: Dict[str, Dict]
) -> None:
    """
    Розраховує продукти для кейтерінгу.
    Формула: вага_інгредієнта × кількість_порцій
    
    ВАЖЛИВО: Використовуємо вагу інгредієнтів з колонки E (сирі інгредієнти ДО приготування),
    а НЕ вихід страви з колонки B (готовий вихід ПІСЛЯ приготування).
    """
    for ingredient in recipe.ingredients:
        if (ingredient.unit or "") == GROUP_UNIT:
            continue
        # ingredient.weight_per_portion - це вага інгредієнта з колонки E (сирі інгредієнти)
        total_weight = ingredient.weight_per_portion * quantity
        product_totals[ingredient.product_name]["quantity"] += total_weight
        product_totals[ingredient.product_name]["unit"] = ingredient.unit or "г"


def _calculate_box_products(
    recipe: models.Recipe, 
    quantity: float, 
    product_totals: Dict[str, Dict]
) -> None:
    """
    Розраховує продукти для боксів.
    Формула: вага_інгредієнта × кількість_компонентів × кількість_порцій
    
    ВАЖЛИВО: Використовуємо вагу інгредієнтів з колонки E (сирі інгредієнти ДО приготування),
    а НЕ вихід страви/компонента з колонки B (готовий вихід ПІСЛЯ приготування).
    """
    for component in recipe.components:
        component_qty = component.quantity_per_portion or 1.0
        
        for ingredient in component.ingredients:
            if (ingredient.unit or "") == GROUP_UNIT:
                continue
            # ingredient.weight_per_unit - це вага інгредієнта з колонки E (сирі інгредієнти)
            # Формула: вага × кількість_компонентів × кількість_порцій
            total_weight = ingredient.weight_per_unit * component_qty * quantity
            product_totals[ingredient.product_name]["quantity"] += total_weight
            product_totals[ingredient.product_name]["unit"] = ingredient.unit or "г"
    
    # Також додаємо прямі інгредієнти (якщо є)
    for ingredient in recipe.ingredients:
        if (ingredient.unit or "") == GROUP_UNIT:
            continue
        # ingredient.weight_per_portion - це вага інгредієнта з колонки E (сирі інгредієнти)
        total_weight = ingredient.weight_per_portion * quantity
        product_totals[ingredient.product_name]["quantity"] += total_weight
        product_totals[ingredient.product_name]["unit"] = ingredient.unit or "г"


# ============ CRUD операції ============

def create_recipe(
    db: Session,
    name: str,
    recipe_type: Literal["catering", "box"],
    category: Optional[str] = None,
    weight_per_portion: Optional[float] = None,
    notes: Optional[str] = None,
    ingredients: Optional[List[Dict]] = None,
    components: Optional[List[Dict]] = None
) -> models.Recipe:
    """
    Створює нову техкарту.
    
    Args:
        db: Сесія бази даних
        name: Назва страви
        recipe_type: Тип техкарти
        category: Категорія
        weight_per_portion: Вага порції
        ingredients: Список інгредієнтів (для кейтерінгу)
        components: Список компонентів з інгредієнтами (для боксів)
    """
    recipe = models.Recipe(
        name=name,
        recipe_type=recipe_type,
        category=category,
        weight_per_portion=weight_per_portion,
        notes=notes,
    )
    
    if ingredients:
        for idx, ing_data in enumerate(ingredients):
            ingredient = models.RecipeIngredient(
                product_name=ing_data["product_name"],
                weight_per_portion=ing_data["weight_per_portion"],
                unit=ing_data.get("unit", "г"),
                order_index=idx
            )
            recipe.ingredients.append(ingredient)
    
    if components and recipe_type == "box":
        for comp_idx, comp_data in enumerate(components):
            component = models.RecipeComponent(
                name=comp_data["name"],
                quantity_per_portion=comp_data.get("quantity_per_portion", 1.0),
                order_index=comp_idx
            )
            
            for ing_idx, ing_data in enumerate(comp_data.get("ingredients", [])):
                ingredient = models.RecipeComponentIngredient(
                    product_name=ing_data["product_name"],
                    weight_per_unit=ing_data["weight_per_unit"],
                    unit=ing_data.get("unit", "г"),
                    order_index=ing_idx
                )
                component.ingredients.append(ingredient)
            
            recipe.components.append(component)
    
    db.add(recipe)
    db.commit()
    db.refresh(recipe)
    return recipe


def delete_recipe(db: Session, recipe_id: int) -> bool:
    """Видаляє техкарту за ID."""
    recipe = db.query(models.Recipe).filter(models.Recipe.id == recipe_id).first()
    if recipe:
        db.delete(recipe)
        db.commit()
        return True
    return False


def update_recipe(
    db: Session,
    recipe_id: int,
    *,
    name: str,
    recipe_type: Literal["catering", "box"],
    category: Optional[str] = None,
    weight_per_portion: Optional[float] = None,
    notes: Optional[str] = None,
    ingredients: Optional[List[Dict]] = None,
    components: Optional[List[Dict]] = None,
) -> Optional[models.Recipe]:
    """
    Оновлює техкарту та повністю замінює її інгредієнти/компоненти.
    Використовується для ручного редагування з фронтенду.
    """
    recipe = get_recipe_by_id(db, recipe_id)
    if not recipe:
        return None

    recipe.name = name
    recipe.category = category
    recipe.weight_per_portion = weight_per_portion
    recipe.recipe_type = recipe_type
    recipe.notes = notes

    # Повністю замінюємо склад: очищаємо старі зв'язки (delete-orphan каскад)
    recipe.ingredients.clear()
    recipe.components.clear()
    db.flush()

    # Додаємо прямі інгредієнти (для catering або як додаткові для box)
    if ingredients:
        for idx, ing_data in enumerate(ingredients):
            if not ing_data.get("product_name"):
                continue
            ingredient = models.RecipeIngredient(
                product_name=clean_product_name(str(ing_data["product_name"]).strip()),
                weight_per_portion=float(ing_data.get("weight_per_portion") or 0),
                unit=ing_data.get("unit", "г") or "г",
                order_index=idx,
            )
            recipe.ingredients.append(ingredient)

    # Додаємо компоненти (тільки для box)
    if components and recipe_type == "box":
        for comp_idx, comp_data in enumerate(components):
            comp_name = (comp_data.get("name") or "").strip()
            if not comp_name:
                continue

            component = models.RecipeComponent(
                name=comp_name,
                quantity_per_portion=float(comp_data.get("quantity_per_portion") or 1.0),
                order_index=comp_idx,
            )

            for ing_idx, ing_data in enumerate(comp_data.get("ingredients", []) or []):
                if not ing_data.get("product_name"):
                    continue
                component_ingredient = models.RecipeComponentIngredient(
                    product_name=clean_product_name(str(ing_data["product_name"]).strip()),
                    weight_per_unit=float(ing_data.get("weight_per_unit") or 0),
                    unit=ing_data.get("unit", "г") or "г",
                    order_index=ing_idx,
                )
                component.ingredients.append(component_ingredient)

            recipe.components.append(component)

    db.commit()
    db.refresh(recipe)
    return recipe
