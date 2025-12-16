import openpyxl
import os
from sys import executable
import re
from openpyxl.styles import NamedStyle, Font
from copy import copy
from typing import List, Tuple, Dict


class BaseExcelHandler:
    """
    Base class for working with Excel files. Contains basic methods for finding file paths and saving changes.
    """

    def __init__(self):
        """
        Initializes the object by obtaining the path to the folder with the executable file and creating a new workbook.
        """
        # Get the path to the folder where the executable file is located
        self.program_folder_path = os.path.dirname(executable)
        self.file_path = 'excel_file.xlsx'
        self.workbook = openpyxl.Workbook()

    @staticmethod
    def find_excel_path(input_file_path: str) -> str:
        """
        Searches for the path to the Excel file in the input path string.

        :param input_file_path: String with the file path.
        :return: Path to the Excel file.
        :raises ValueError: If the path to the Excel file is not found.
        """
        pattern = r"\w:\\.+\.xlsx?"
        match = re.findall(pattern, input_file_path)
        if match:
            return match[0]
        else:
            raise ValueError("Не найдено соответствий для пути к файлу Excel.")

    def save(self) -> str:
        """
        Saves changes to Excel files, adding "_edited" to the filename.

        :return: Path to the created file.
        """
        # Get the original file name
        file_name, file_extension = os.path.splitext(
            os.path.basename(self.file_path))

        # Create a name for the new file
        new_file_name = f"{file_name}_edited{file_extension}"
        new_file_path = os.path.join(self.program_folder_path, new_file_name)

        # Save the new file
        self.workbook.save(new_file_path)
        print(f"Новый файл сохранен по адресу: {new_file_path}")

        return new_file_path

    def find_customer_name(self, sheet) -> str:
        """
        Searches for the customer's name on the 2nd sheet of the client file.

        :param sheet: Excel sheet to search.
        :return: Customer's name.
        """
        max_row = sheet.max_row
        for row in range(1, max_row + 1):
            if sheet.cell(row=row, column=1).value == 'Замовник:':
                customer_name = sheet.cell(row=row, column=2).value
                return customer_name
        return "нет имени"  # "no name"


class ExcelHandlerForBoxMenu(BaseExcelHandler):
    """
    Class for processing menu in the Excel file.
    """

    def __init__(self, menu_file_path: str):
        """
        Initializes the object with the file path and loads the first two sheets.

        :param menu_file_path: Path to the Excel file.
        """
        super().__init__()  # Call the base class initializer
        self.menu_file_path = self.find_excel_path(menu_file_path)
        self.workbook_menu = openpyxl.load_workbook(self.menu_file_path)
        self.wb_sheet1 = self.workbook_menu.worksheets[0]
        self.wb_sheet2 = self.workbook_menu.worksheets[1]

    def handle_sheet_1(self) -> None:
        """
        Processes the first sheet:
        - Removes merged cells.
        - Deletes images.
        - Deletes the first row.
        - Deletes rows and columns according to certain conditions.
        - Creates a header.
        """
        # Remove merged cells
        merged_cells_ranges = self.wb_sheet1.merged_cells.ranges
        if merged_cells_ranges:
            for merged_cell in list(merged_cells_ranges):
                self.wb_sheet1.unmerge_cells(str(merged_cell))

        # Remove images
        for image in self.wb_sheet1._images[:]:
            self.wb_sheet1._images.remove(image)

        # Delete the first row
        self.wb_sheet1.delete_rows(1, 1)

        # Delete rows starting with "Обладнання та сервіс"
        max_row = self.wb_sheet1.max_row
        for row in range(1, max_row + 1):
            if self.wb_sheet1.cell(row=row, column=2).value == 'Обладнання та сервіс':
                break

        self.wb_sheet1.delete_rows(row, max_row - row + 1)
        self.wb_sheet1.delete_cols(5, 5)

        customer = self.find_customer_name(self.wb_sheet2)
        # Create a header
        self.wb_sheet1.insert_rows(1, amount=6)
        for row in range(1, self.wb_sheet2.max_row + 1):
            match self.wb_sheet2.cell(row=row, column=1).value:
                case 'Дата проведення:':
                    self.wb_sheet1.cell(row=1, column=1).value = 'Дата'
                    cell = self.wb_sheet1.cell(row=1, column=3)
                    cell.value = self.wb_sheet2.cell(row=row, column=2).value
                    date_style = NamedStyle(
                        name='short_date', number_format='DD.MM.YYYY')
                    cell.style = date_style
                    cell.font = Font(size=16)
                case 'Місце проведення:':
                    self.wb_sheet1.cell(row=3, column=1).value = 'Локація'
                    self.wb_sheet1.cell(row=3, column=2).value = self.wb_sheet2.cell(
                        row=row, column=2).value
                case 'Час:':
                    cell = self.wb_sheet1.cell(row=2, column=3)
                    cell.value = self.wb_sheet2.cell(row=row, column=2).value
                    time_style = NamedStyle(
                        name='short_time',
                        number_format='HH:MM'
                    )
                    cell.style = time_style
                    cell.font = Font(size=16)
                case 'Назва проекту (привід)':
                    self.wb_sheet1.cell(row=4, column=1).value = 'Формат'
                    self.wb_sheet1.cell(row=4, column=2).value = self.wb_sheet2.cell(
                        row=row, column=2).value
                case 'Кількість осіб:':
                    break

        self.wb_sheet1.cell(row=2, column=1).value = 'Замовник'
        self.wb_sheet1.cell(row=2, column=2).value = customer
        self.wb_sheet1.cell(row=5, column=1).value = 'Коментар'


class ExcelHandlerForBox(BaseExcelHandler):
    """
    Class for processing specific order lists in the Excel file.
    """

    def __init__(self, purchase_file_path: str):
        """
        Initializes the object with the file path and loads the first sheet.

        :param purchase_file_path: Path to the Excel file.
        """
        super().__init__()  # Call the base class initializer
        self.file_path = self.find_excel_path(purchase_file_path)
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.wb_sheet1 = self.workbook.worksheets[0]
        self.label = 'point'
        self.label_in_box = 'in'

    def add_menus_to_purchase_file(self, customer_name: str, menu: openpyxl.Workbook) -> None:
        """
        Creates a menu with the customer's name in the calculations file.

        :param customer_name: Customer's name.
        :param menu: Workbook with the menu.
        """
        wb_sheet1 = menu.worksheets[0]

        self.workbook.create_sheet(title=f'Меню {customer_name}')
        wb_sheet_menu = self.workbook[f'Меню {customer_name}']

        # Copy data row by row and column by column
        for row in wb_sheet1.iter_rows():
            for cell in row[:4]:
                new_cell = wb_sheet_menu.cell(
                    row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Copy column widths (limit to the first 4 columns)
        for col_letter in list(wb_sheet1.column_dimensions.keys())[:4]:
            wb_sheet_menu.column_dimensions[col_letter].width = wb_sheet1.column_dimensions[col_letter].width

        # Copy row heights
        for row in wb_sheet1.row_dimensions:
            wb_sheet_menu.row_dimensions[row].height = wb_sheet1.row_dimensions[row].height

        # Merge cells in the new sheet
        wb_sheet_menu.merge_cells('C1:D1')
        wb_sheet_menu.merge_cells('C2:D5')


        # Fix the row with "№ п/п" and the next row
        max_row = wb_sheet_menu.max_row
        for row in range(1, max_row + 1):
            if wb_sheet_menu.cell(row=row, column=1).value == '№ п/п':
                wb_sheet_menu.row_dimensions[row].height = 50
                wb_sheet_menu.row_dimensions[row + 1].height = 15

        wb_sheet_menu.row_dimensions[1].height = 30
        wb_sheet_menu.row_dimensions[2].height = 15
        wb_sheet_menu.column_dimensions['A'].width = 10


    def create_dish_list_wb1_sheet1(self, menu_file_path: str) -> Tuple[str, Dict[str, int]]:
        """
        Creates a list of dishes based on the first order sheet.

        :param menu_file_path: Path to the Excel file.
        :return: Name of the client and a dictionary with the names and quantities of dishes.
        """
        self.menu_file_path = self.find_excel_path(menu_file_path)
        self.menu_wb = openpyxl.load_workbook(
            self.menu_file_path, data_only=True)
        wb_sheet1 = self.menu_wb.worksheets[0]
        wb_sheet2 = self.menu_wb.worksheets[1]
        customer_name = self.find_customer_name(wb_sheet2)
        max_row = wb_sheet1.max_row
        dish_dict = dict()
        for row in range(1, max_row + 1):
            col1_value = wb_sheet1.cell(row=row, column=1).value
            col2_value = wb_sheet1.cell(row=row, column=2).value
            col4_value = wb_sheet1.cell(row=row, column=4).value
            if col2_value != 'Обладнання та сервіс':
                if col1_value not in [None, ''] and isinstance(col1_value, int) or \
                        (isinstance(col1_value, str) and col1_value.isdigit()):
                    if col2_value in dish_dict.keys():
                        dish_dict[col2_value] += col4_value
                    else:
                        dish_dict[col2_value] = col4_value
            else:
                break
        return customer_name, dish_dict

    def dish_finder_calc(self, dish_dict: Dict[str, int]) -> List[Tuple[int, int, str]]:
        """
        Finds and counts the quantity of dishes based on the dish list.

        :param dish_dict: Dictionary with dish names and their quantities.
        :return: List of technical cards.
        """
        max_row = self.wb_sheet1.max_row
        tech_cards = []
        for row in range(1, max_row + 1):
            col3_value = self.wb_sheet1.cell(row=row, column=3).value
            if col3_value in dish_dict.keys() and self.wb_sheet1.cell(row=row, column=7).value == self.label:
                col1 = self.wb_sheet1.cell(row=row, column=1)
                if col1.value in ['', None]:
                    col1.value = dish_dict[col3_value]
                else:
                    col1.value += dish_dict[col3_value]
                for i in range(row + 1, max_row + 1):
                    cell = self.wb_sheet1.cell(row=i, column=7).value
                    if cell == self.label:
                        tech_cards.append((row, i - 1, col3_value))
                        break
        return tech_cards

    def create_tech_card_sheet(self, tech_cards: List[Tuple[int, int, str]], customer: str, dish_dict: Dict[str, int]) -> None:
        """
        Creates a list of technical cards based on found technical cards and the dish list.

        :param tech_cards: List of technical cards.
        :param customer: Customer's name.
        :param dish_dict: Dictionary with dish names and their quantities.
        """
        self.workbook.create_sheet(title=f'техкарта {customer}')
        wb_sheet_tech_card = self.workbook[f'техкарта {customer}']
        # Copying rows of technical cards
        idx_max_row_in_tech_card = 1
        for tech_card in tech_cards:
            for row in range(tech_card[0], tech_card[1] + 1):
                for col in range(1, 8):  # всего 7 столбцов +1 для range
                    new_cell = wb_sheet_tech_card.cell(
                        row=idx_max_row_in_tech_card, column=col)
                    cell = self.wb_sheet1.cell(row=row, column=col)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                idx_max_row_in_tech_card += 1
        wb_sheet_tech_card.cell(
            row=idx_max_row_in_tech_card, column=7).value = self.label
        # Copying column widths
        for col in self.wb_sheet1.column_dimensions:
            wb_sheet_tech_card.column_dimensions[col].width = self.wb_sheet1.column_dimensions[col].width

        # Setting values in the technical card
        max_row = wb_sheet_tech_card.max_row
        label_row = 1
        label_in_box_row = 0
        for row in range(1, max_row + 1):
            col3_value = wb_sheet_tech_card.cell(row=row, column=3).value
            if col3_value in dish_dict.keys() and wb_sheet_tech_card.cell(row=row, column=7).value == self.label:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = dish_dict[col3_value]
                label_row = row
                label_in_box_row = 0
                wb_sheet_tech_card.cell(
                    row=row, column=5).value = ''
                continue
            if wb_sheet_tech_card.cell(row=row, column=7).value == self.label_in_box:
                wb_sheet_tech_card.cell(
                    row=row, column=5).value = ''
                label_in_box_row = row
                continue
            if label_in_box_row:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = f'=E{row}*A{label_row}*A{label_in_box_row}'
            else:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = f'=E{row}*A{label_row}'
                
        # Adjusting column sizes
        wb_sheet_tech_card.column_dimensions['C'].width = 80
        wb_sheet_tech_card.column_dimensions['D'].width = 50

        wb_sheet_tech_card.delete_cols(6, 2)
