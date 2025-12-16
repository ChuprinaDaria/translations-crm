import openpyxl
import os
from sys import executable
import re
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
from copy import copy


class Base_ExcelHandler:
    """
    Базовый класс для работы с Excel файлами. Содержит основные методы для поиска путей к файлам и сохранения изменений.
    """

    def __init__(self):
        # Получаем путь к папке, в которой находится исполняемый файл
        # Получаем путь к исполняемому файлу .exe
        self.program_folder_path = os.path.dirname(executable)
        self.file_path = 'excel_file.xlsx'
        self.workbook = openpyxl.Workbook()

    @staticmethod
    def find_excel_path(input_file_path: str) -> str:
        """
        Ищет путь к файлу Excel в строке пути.

        :param input_file_path: Строка с путем к файлу.
        :return: Путь к файлу Excel.
        :raises ValueError: Если путь к файлу Excel не найден.
        """
        pattern = r"\w:\\.+\.xlsx?"
        match = re.findall(pattern, input_file_path)
        if match:
            return match[0]
        else:
            raise ValueError("Не найдено соответствий для пути к файлу Excel.")

    def saver(self) -> None:
        """
        Сохраняет изменения в Excel файлах, добавляя к имени "_edited".
        :return: путь к созданому файлу.
        """
        # Получаем имя исходного файла
        file_name, file_extension = os.path.splitext(
            os.path.basename(self.file_path))

        # Создаем имя для нового файла
        new_file_name = f"{file_name}_edited{file_extension}"
        new_file_path = os.path.join(self.program_folder_path, new_file_name)

        # Сохраняем новый файл
        self.workbook.save(new_file_path)
        print(f"Новый файл сохранен по адресу: {new_file_path}")

        return new_file_path


class ExcelHandler_for_order_list(Base_ExcelHandler):
    """
    Класс для обработки конкретного списка заказов в файле Excel.
    """

    def __init__(self, file_path: str):
        """
        Инициализирует объект с путем к файлу и загружает первый три листа.

        :param file_path: Путь к файлу Excel.
        """
        super().__init__()
        self.file_path = self.find_excel_path(file_path)
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.sheet1 = self.workbook.worksheets[0]
        self.sheet2 = self.workbook.worksheets[1]
        self.sheet3 = self.workbook.worksheets[2]

    def handler_for_sheet_1(self) -> None:
        """
        Обрабатывает первый лист:
        - Убирает объединение ячеек.
        - Удаляет изображения.
        - Удаляет первую строку.
        - Удаляет строки и колонки согласно определенным условиям.
        - Создает заголовок.
        """
        # Убираем объединение ячеек
        merged_cells_ranges = self.sheet1.merged_cells.ranges
        if merged_cells_ranges:
            for merged_cell in list(merged_cells_ranges):
                self.sheet1.unmerge_cells(str(merged_cell))

        # Удаление изображений
        # Используем [:], чтобы сделать копию списка
        for image in self.sheet1._images[:]:
            self.sheet1._images.remove(image)

        # Удаляем первую строку
        self.sheet1.delete_rows(1, 1)

        # Удаляем строки начиная с "Обладнання та сервіс"
        max_row = self.sheet1.max_row
        for row in range(1, max_row + 1):
            if self.sheet1.cell(row=row, column=2).value == 'Обладнання та сервіс':
                break
            # Поправляем строку с "№ п/п" и следующую
            elif self.sheet1.cell(row=row, column=1).value == '№ п/п':
                self.sheet1.row_dimensions[row].height = 55
                self.sheet1.row_dimensions[row + 1].height = 15

        # Удаление лишних колонок и строк
        self.sheet1.delete_rows(row, max_row - row + 1)
        self.sheet1.delete_cols(5, 5)

        # Поправляем строку с "№ п/п" и следующую
        max_row = self.sheet1.max_row
        for row in range(1, max_row + 1):
            if self.sheet1.cell(row=row, column=1).value == '№ п/п':
                self.sheet1.row_dimensions[row].height = 55
                self.sheet1.row_dimensions[row + 1].height = 15
        # Поправляем колонку
        self.sheet1.column_dimensions['A'].width = 10

    def handler_for_sheet_2(self) -> None:
        """
        Обрабатывает второй лист:
        - Удаляет изображения.
        - Удаляет строки и колонки согласно определенным условиям.
        """
        # Удаление изображений
        # Используем [:], чтобы сделать копию списка
        for image in self.sheet2._images[:]:
            self.sheet2._images.remove(image)

        # Получаем все объединенные ячейки
        merged_cells = list(self.sheet2.merged_cells.ranges)

        max_row = self.sheet2.max_row
        for row in range(1, max_row + 1):
            if self.sheet2.cell(row=row, column=2).value == 'Одноразовая посуда':
                for merged_cell in merged_cells:
                    if merged_cell.min_row == row:
                        try:
                            self.sheet2.unmerge_cells(str(merged_cell))
                        except:
                            print(f"{merged_cell}\n")

        for merged_cell in merged_cells:
            if merged_cell.min_row == 1:
                try:
                    self.sheet2.unmerge_cells(str(merged_cell))
                except:
                    print(f"{merged_cell}\n")

        # Удаление лишних колонок и строк
        # Удаляем первую строку
        self.sheet2.delete_rows(1, 1)
        max_row = self.sheet2.max_row
        for row in range(max_row + 1, 1, -1):
            # Удаляем строки начиная с пустой клетки второго столбца, идя с конца
            if self.sheet2.cell(row=row, column=2).value not in ['', None]:
                self.sheet2.delete_rows(row + 1, max_row - row + 1)
                break

        self.sheet2.delete_cols(7, 2)
        self.sheet2.delete_cols(3, 3)

        # Поправляем первую строку
        self.sheet2.row_dimensions[1].height = 35

    def handler_for_sheet_3(self) -> None:
        """
        Обрабатывает третий лист:
        - Удаляет изображения.
        - Удаляет строки начиная с "Обладнання та сервіс".
        - Удаляет первую строку.
        """
        # Удаление изображений
        # Используем [:], чтобы сделать копию списка
        for image in self.sheet3._images[:]:
            self.sheet3._images.remove(image)

        # Удаляем строки начиная с "Обладнання та сервіс"
        max_row = self.sheet3.max_row
        del_list_for_rows = []
        for row in range(1, max_row + 1):
            if self.sheet3.cell(row=row, column=1).value == 'Кухня':
                del_list_for_rows.append(row)
            if self.sheet3.cell(row=row, column=1).value == 'Обслуговування':
                del_list_for_rows.append(row)
            if self.sheet3.cell(row=row, column=1).value == 'Транспортні витрати':
                del_list_for_rows.append(row)
                break

        self.sheet3.delete_rows(del_list_for_rows[2], max_row - row + 1)
        self.sheet3.delete_rows(
            del_list_for_rows[0], del_list_for_rows[1] - del_list_for_rows[0])

        # Удаляем первую строку
        self.sheet3.delete_rows(1, 1)


class ExcelHandler_for_purchase_list(Base_ExcelHandler):
    """
    Класс для обработки списка покупок в файле Excel.
    """

    def __init__(self, order_file_path_list: str, dish_file_path: str):
        """
        Инициализирует объект с путями к файлам заказов и блюд, и загружает соответствующие листы.

        :param order_file_path: Путь к файлу с заказами.
        :param dish_file_path: Путь к файлу с блюдами.
        """
        super().__init__()
        self.file_path = self.find_excel_path(dish_file_path)
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.wb_sheet1 = self.workbook.worksheets[0]

        self.order_file_path_list = order_file_path_list

        self.label = 'point'

    def add_menu_from_order_file(self, path_to_wb: str) -> None:
        """
        Создает меню с именем заказчика в файле с расчетами.

        :param path_to_wb: путь к файлу excel с обрезаной таблецей клиенского заказа
        """
        customer = self.find_customer_name(path_to_wb)
        self.workbook.create_sheet(title=f'Меню {customer}')
        wb_sheet_menu = self.workbook[f'Меню {customer}']

        wb = openpyxl.load_workbook(path_to_wb, data_only=True)
        wb_sheet3 = wb.worksheets[2]

        path_to_wb = self.find_excel_path(path_to_wb)
        wb = openpyxl.load_workbook(path_to_wb, data_only=True)
        order_file_sheet3 = wb.worksheets[0]

        # Копируем данные построчно и поколоночно
        for row in order_file_sheet3.iter_rows():
            for cell in row[:4]:
                # Копируем стили ячеек
                new_cell = wb_sheet_menu.cell(
                    row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Копирование ширины колонок (ограничиваем до первых 4 колонок)
        for col_letter in list(order_file_sheet3.column_dimensions.keys())[:4]:
            wb_sheet_menu.column_dimensions[col_letter].width = order_file_sheet3.column_dimensions[col_letter].width

        # Копирование высоты строк
        for row in order_file_sheet3.row_dimensions:
            wb_sheet_menu.row_dimensions[row].height = order_file_sheet3.row_dimensions[row].height

        # Создаем заголовок (шапку)
        wb_sheet_menu.insert_rows(1, amount=6)
        wb_sheet_menu.merge_cells('C1:D1')
        wb_sheet_menu.merge_cells('C2:D5')
        max_row = wb_sheet3.max_row
        for row in range(1, max_row + 1):
            match wb_sheet3.cell(row=row, column=1).value:
                case 'Дата проведення:':
                    wb_sheet_menu.cell(row=1, column=1).value = 'Дата'
                    cell = wb_sheet_menu.cell(row=1, column=3)
                    cell.value = wb_sheet3.cell(row=row, column=2).value
                    date_style = NamedStyle(
                        name='short_date', number_format='DD.MM.YYYY')
                    cell.style = date_style
                    # Устанавливаем размер шрифта на 14
                    cell.font = Font(size=16)
                case 'Місце проведення:':
                    wb_sheet_menu.cell(row=3, column=1).value = 'Локація'
                    wb_sheet_menu.cell(row=3, column=2).value = wb_sheet3.cell(
                        row=row, column=2).value
                case 'Час:':
                    cell = wb_sheet_menu.cell(row=2, column=3)
                    cell.value = wb_sheet3.cell(row=row, column=2).value
                    time_style = NamedStyle(
                        name='short_time',
                        number_format='HH:MM'
                    )
                    cell.style = time_style
                    cell.font = Font(size=16)
                case 'Назва проекту (привід)':
                    wb_sheet_menu.cell(row=4, column=1).value = 'Формат'
                    wb_sheet_menu.cell(row=4, column=2).value = wb_sheet3.cell(
                        row=row, column=2).value
                case 'Кухня':
                    break
        wb_sheet_menu.cell(row=2, column=1).value = 'Замовник'
        wb_sheet_menu.cell(row=2, column=2).value = customer
        wb_sheet_menu.cell(row=5, column=1).value = 'Кометар'

        # Поправляем строку с "№ п/п" и следующую
        max_row = wb_sheet_menu.max_row
        for row in range(1, max_row + 1):
            if wb_sheet_menu.cell(row=row, column=1).value == '№ п/п':
                wb_sheet_menu.row_dimensions[row].height = 55
                wb_sheet_menu.row_dimensions[row + 1].height = 15
        # Поправляем колонку
        wb_sheet_menu.column_dimensions['A'].width = 10

    def create_dish_list_sheet1(self, path_to_wb: str) -> dict:
        """
        Создает список блюд на основе первого листа заказов.

        :param path_to_wb: путь к файлу excel с таблецей клиенского заказа
        :return: Словарь с названиями блюд и их количеством.
        """
        path_to_wb = self.find_excel_path(path_to_wb)
        wb = openpyxl.load_workbook(path_to_wb, data_only=True)
        wb_sheet1 = wb.worksheets[0]
        max_row = wb_sheet1.max_row
        dish_dict = dict()
        for row in range(1, max_row + 1):
            col1_value = wb_sheet1.cell(row=row, column=1).value
            col2_value = wb_sheet1.cell(row=row, column=2).value
            col4_value = wb_sheet1.cell(row=row, column=4).value
            if col2_value != 'Обладнання та сервіс':
                if col2_value not in ['', None] and col1_value not in ['', None] and isinstance(col1_value, int):
                    if col2_value in dish_dict.keys():
                        dish_dict.update(
                            {col2_value: dish_dict[col2_value] + col4_value})
                    else:
                        dish_dict.update({col2_value: col4_value})
            else:
                break
        return dish_dict

    def dish_finder_calc(self, dish_dict: dict) -> list:
        """
        Находит и считает количество блюд на основе списка блюд.

        :param dish_dict: Словарь с названиями блюд и их количеством.
        :return: Список техкарт.
        """
        max_row = self.wb_sheet1.max_row
        tech_cards = []
        for row in range(1, max_row + 1):
            col3_value = self.wb_sheet1.cell(row=row, column=3).value
            if col3_value in dish_dict.keys() and self.wb_sheet1.cell(row=row, column=7).value == self.label:
                for i in range(row + 1, max_row + 1):
                    cell = self.wb_sheet1.cell(row=i, column=7).value
                    if cell == self.label:
                        tech_cards.append([row, i-1, col3_value])
                        break
                if self.wb_sheet1.cell(row=row, column=1).value in ['', None]:
                    self.wb_sheet1.cell(
                        row=row, column=1).value = dish_dict[col3_value]
                else:
                    self.wb_sheet1.cell(
                        row=row, column=1).value += dish_dict[col3_value]
        return tech_cards

    def find_customer_name(self, path_to_wb: str) -> str:
        """
        Ищет имя заказчика на 3 листе клиенского файла.

        :param path_to_wb: путь к файлу excel таблецей клиенского заказа.
        :return: имя заказчика.
        """
        path_to_wb = self.find_excel_path(path_to_wb)
        wb = openpyxl.load_workbook(path_to_wb, data_only=True)
        order_file_sheet3 = wb.worksheets[2]

        max_row = order_file_sheet3.max_row
        for row in range(1, max_row + 1):
            if order_file_sheet3.cell(row=row, column=1).value == 'Замовник:':
                customer_name = order_file_sheet3.cell(row=row, column=2).value
                break
        return customer_name

    def create_tech_card_sheet(self, tech_cards: list, dish_dict: dict, path_to_wb: str) -> None:
        """
        Создает список техкарт на основе найденных техкарт и списка блюд.

        :param tech_cards: Список техкарт.
        :param dish_dict: Словарь с названиями блюд и их количеством.
        :param path_to_wb: путь к файлу excel с таблецей клиенского заказа
        """
        customer = self.find_customer_name(path_to_wb)
        self.workbook.create_sheet(title=f'техкарта {customer}')
        wb_sheet_tech_card = self.workbook[f'техкарта {customer}']
        # Копирование строк меню
        idx_max_row_in_tech_card = 1
        for tech_card in tech_cards:
            for row in range(tech_card[0], tech_card[1] + 1):
                for col in range(1, 8):  # всего 7 столбцов +1 для range
                    new_cell = wb_sheet_tech_card.cell(
                        row=idx_max_row_in_tech_card, column=col)
                    cell = self.wb_sheet1.cell(row=row, column=col)
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                idx_max_row_in_tech_card += 1
        wb_sheet_tech_card.cell(
            row=idx_max_row_in_tech_card, column=7).value = self.label
        # Копирование ширины колонок
        for col in self.wb_sheet1.column_dimensions:
            wb_sheet_tech_card.column_dimensions[col].width = self.wb_sheet1.column_dimensions[col].width

        # Setting values in the technical card
        max_row = wb_sheet_tech_card.max_row
        label_row = 1
        label_in_box_row = 0
        for row in range(1, max_row + 1):
            col3_value = wb_sheet_tech_card.cell(row=row, column=3).value
            if col3_value in dish_dict.keys() and wb_sheet_tech_card.cell(row=row, column=7).value == self.label:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = dish_dict[col3_value]
                label_row = row
                label_in_box_row = 0
                wb_sheet_tech_card.cell(
                    row=row, column=5).value = ''
                continue
            if wb_sheet_tech_card.cell(row=row, column=7).value == self.label:
                wb_sheet_tech_card.cell(
                    row=row, column=5).value = ''
                label_in_box_row = row
                continue
            if label_in_box_row:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = f'=E{row}*A{label_row}*A{label_in_box_row}'
            else:
                wb_sheet_tech_card.cell(
                    row=row, column=1).value = f'=E{row}*A{label_row}'

        # поправляем размер столбцов
        wb_sheet_tech_card.column_dimensions['C'].width = 80
        wb_sheet_tech_card.column_dimensions['D'].width = 50

        wb_sheet_tech_card.delete_cols(6, 2)
