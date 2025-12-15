"""
Сервіс для генерації Excel файлів закупки.

Підтримує два типи розрахунків:
- catering: кейтерінг (проста структура)
- box: бокси (трирівнева структура)
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple, Iterable, Optional, Literal

from sqlalchemy.orm import Session, joinedload

import models

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise RuntimeError("openpyxl is required for purchase Excel export") from e


# ============ Стилі для Excel ============

HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FILL_BOX = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def _style_header_row(ws, row: int, fill=None, num_cols: int = 7):
    """Застосовує стиль до заголовка."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill or HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN


def _auto_column_width(ws, min_width: int = 10, max_width: int = 50):
    """Автоматично підлаштовує ширину колонок."""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width


def _load_kps_with_items(db: Session, kp_ids: List[int]) -> List[models.KP]:
    """Завантажує КП разом із позиціями меню, стравами та клієнтом."""
    return (
        db.query(models.KP)
        .options(
            joinedload(models.KP.items).joinedload(models.KPItem.item),
            joinedload(models.KP.client),
        )
        .filter(models.KP.id.in_(kp_ids))
        .all()
    )


def _iter_kp_items(kp: models.KP) -> Iterable[Tuple[int, str, float, str | None, float | None, float | None, float]]:
    """
    Повертає ітератор по позиціях КП у вигляді:
    (№, назва, кількість, одиниця, вага, ціна, сума)
    """
    index = 1
    for kp_item in kp.items:
        if kp_item.item and kp_item.item.name:
            name = kp_item.item.name
        elif kp_item.name:
            name = kp_item.name
        else:
            name = "Без назви"

        qty = kp_item.quantity or 0

        unit = None
        if kp_item.item and kp_item.item.unit:
            unit = kp_item.item.unit
        elif kp_item.unit:
            unit = kp_item.unit

        weight = None
        if kp_item.item and kp_item.item.weight is not None:
            weight = float(kp_item.item.weight)
        elif kp_item.weight is not None:
            weight = float(kp_item.weight)

        price = None
        if kp_item.item and kp_item.item.price is not None:
            price = float(kp_item.item.price)
        elif kp_item.price is not None:
            price = float(kp_item.price)

        total = (price or 0.0) * qty

        yield index, name, qty, unit, weight, price, total
        index += 1


def _safe_sheet_title(base: str) -> str:
    """Робить безпечну назву для аркуша Excel (макс. 31 символ)."""
    invalid_chars = set(r"[]:*?/\\")
    cleaned = "".join(c for c in base if c not in invalid_chars)
    if not cleaned:
        cleaned = "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def _build_dish_totals(kps: List[models.KP]) -> Dict[str, Dict[str, float | str | None]]:
    """Формує агрегований словник страв."""
    dish_totals: Dict[str, Dict[str, float | str | None]] = defaultdict(
        lambda: {"quantity": 0, "unit": None}
    )

    for kp in kps:
        for kp_item in kp.items:
            name = None
            if kp_item.item and kp_item.item.name:
                name = kp_item.item.name
            elif kp_item.name:
                name = kp_item.name
            else:
                name = "Без назви"

            qty = kp_item.quantity or 0

            unit = None
            if kp_item.item and kp_item.item.unit:
                unit = kp_item.item.unit
            elif kp_item.unit:
                unit = kp_item.unit

            entry = dish_totals[name]
            entry["quantity"] = (entry["quantity"] or 0) + qty
            if entry["unit"] is None and unit:
                entry["unit"] = unit

    return dish_totals


def _determine_calculation_type(kps: List[models.KP]) -> str:
    """Визначає тип розрахунку на основі event_group КП."""
    for kp in kps:
        if kp.event_group and "бокс" in kp.event_group.lower():
            return "box"
    return "catering"


def generate_purchase_excel(
    db: Session, 
    kp_ids: List[int],
    calculation_type: Optional[Literal["catering", "box", "auto"]] = "auto"
) -> Tuple[bytes, str]:
    """
    Генерує Excel-файл для закупки на основі вибраних КП.

    Args:
        db: Сесія бази даних
        kp_ids: Список ID КП
        calculation_type: Тип розрахунку ('catering', 'box', 'auto')

    Повертає (bytes, filename).
    """
    if not kp_ids:
        raise ValueError("Список KP ID порожній")

    kps = _load_kps_with_items(db, kp_ids)
    if not kps:
        raise ValueError("Не знайдено жодного КП для вказаних ID")

    # Визначаємо тип розрахунку
    if calculation_type == "auto":
        calculation_type = _determine_calculation_type(kps)

    dish_totals = _build_dish_totals(kps)

    wb = Workbook()
    header_fill = HEADER_FILL_BOX if calculation_type == "box" else HEADER_FILL

    # ---------- Лист "Закупка" ----------
    ws_purchase = wb.active
    ws_purchase.title = "Закупка"
    
    # Заголовок з типом
    type_label = "🍽️ Кейтерінг" if calculation_type == "catering" else "📦 Бокси"
    ws_purchase["A1"] = f"Закупка продуктів ({type_label})"
    ws_purchase["A1"].font = Font(bold=True, size=14)
    ws_purchase.merge_cells("A1:C1")
    
    ws_purchase.append(["Страва", "Одиниця", "Кількість"])
    _style_header_row(ws_purchase, 2, header_fill, 3)

    for name, data in sorted(dish_totals.items(), key=lambda x: str(x[0])):
        ws_purchase.append([
            name,
            data.get("unit") or "",
            data.get("quantity") or 0,
        ])

    _auto_column_width(ws_purchase)

    # ---------- Лист "КП" ----------
    ws_kps = wb.create_sheet("КП")
    ws_kps.append([
        "KP ID", "Назва КП", "Клієнт", "Тип", "Дата події",
        "Локація", "К-сть гостей", "Статус", "Сума загальна"
    ])
    _style_header_row(ws_kps, 1, header_fill, 9)

    for kp in kps:
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        event_date_str = ""
        if kp.event_date:
            if isinstance(kp.event_date, datetime):
                event_date_str = kp.event_date.strftime("%Y-%m-%d")
            else:
                event_date_str = str(kp.event_date)

        ws_kps.append([
            kp.id,
            kp.title,
            client_name,
            kp.event_group or "",
            event_date_str,
            kp.event_location or "",
            kp.people_count or 0,
            kp.status or "",
            float(kp.total_amount or 0),
        ])

    _auto_column_width(ws_kps)

    # ---------- Листи "Меню {клієнт}" ----------
    for kp in kps:
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        menu_title = client_name or kp.title or f"KP {kp.id}"
        ws_menu = wb.create_sheet(_safe_sheet_title(f"Меню {menu_title}"))

        # Шапка
        ws_menu["A1"] = "Дата проведення:"
        if kp.event_date:
            if isinstance(kp.event_date, datetime):
                ws_menu["B1"] = kp.event_date.strftime("%d.%m.%Y")
            else:
                ws_menu["B1"] = str(kp.event_date)

        ws_menu["A2"] = "Замовник:"
        ws_menu["B2"] = client_name

        ws_menu["A3"] = "Місце проведення:"
        ws_menu["B3"] = kp.event_location or ""

        ws_menu["A4"] = "Формат:"
        ws_menu["B4"] = kp.event_format or kp.title or ""

        ws_menu["A5"] = "Кількість осіб:"
        ws_menu["B5"] = kp.people_count or 0

        ws_menu["A6"] = "Тип:"
        ws_menu["B6"] = kp.event_group or ""

        # Заголовок таблиці
        ws_menu.append([])
        headers = ["№ п/п", "Назва страви", "Кількість", "Одиниця", "Вага, г", "Ціна, грн", "Сума, грн"]
        ws_menu.append(headers)
        _style_header_row(ws_menu, 8, header_fill, 7)

        for index, name, qty, unit, weight, price, total in _iter_kp_items(kp):
            ws_menu.append([
                index, name, qty, unit or "", weight or "", price or "", total
            ])

        _auto_column_width(ws_menu)

    # ---------- Лист "Продукти" з техкартами ----------
    try:
        from recipe_service import calculate_purchase_from_kps
        
        purchase_result = calculate_purchase_from_kps(db, kp_ids, calculation_type)
        products = purchase_result.get("products", {})
        dishes_without_recipe = purchase_result.get("dishes_without_recipe", [])
        calc_type = purchase_result.get("calculation_type", calculation_type)
        
        if products:
            ws_products = wb.create_sheet("Продукти")
            
            # Заголовок з типом розрахунку
            type_label = "🍽️ Кейтерінг" if calc_type == "catering" else "📦 Бокси"
            ws_products["A1"] = f"Розрахунок продуктів ({type_label})"
            ws_products["A1"].font = Font(bold=True, size=14)
            ws_products.merge_cells("A1:C1")
            
            ws_products.append(["Продукт", "Кількість (г)", "Одиниця"])
            _style_header_row(ws_products, 2, header_fill, 3)
            
            for product_name, data in sorted(products.items(), key=lambda x: str(x[0])):
                ws_products.append([
                    product_name,
                    round(data.get("quantity", 0), 2),
                    data.get("unit", "г")
                ])
            
            # Статистика
            ws_products.append([])
            ws_products.append(["Статистика:"])
            ws_products.append(["Страв знайдено:", purchase_result.get("dishes_with_recipe", 0)])
            ws_products.append(["Страв без техкарт:", len(dishes_without_recipe)])
            
            if dishes_without_recipe:
                ws_products.append([])
                ws_products.append(["⚠️ Страви без техкарт:"])
                for dish in dishes_without_recipe:
                    ws_products.append([f"  • {dish}"])
            
            _auto_column_width(ws_products)
            
    except ImportError:
        pass
    except Exception as e:
        print(f"Помилка розрахунку продуктів: {e}")

    # ---------- Листи "техкарта {клієнт}" ----------
    for kp in kps:
        client_name = kp.client_name or (kp.client.name if kp.client else "")
        tech_title = client_name or kp.title or f"KP {kp.id}"
        ws_tech = wb.create_sheet(_safe_sheet_title(f"техкарта {tech_title}"))

        ws_tech.append(["Назва страви", "Кількість", "Одиниця"])
        _style_header_row(ws_tech, 1, header_fill, 3)

        local_totals: Dict[str, Dict[str, float | str | None]] = defaultdict(
            lambda: {"quantity": 0, "unit": None}
        )
        for _idx, name, qty, unit, _weight, _price, _total in _iter_kp_items(kp):
            entry = local_totals[name]
            entry["quantity"] = (entry["quantity"] or 0) + qty
            if entry["unit"] is None and unit:
                entry["unit"] = unit

        for name, data in sorted(local_totals.items(), key=lambda x: str(x[0])):
            ws_tech.append([
                name,
                data.get("quantity") or 0,
                data.get("unit") or "",
            ])

        _auto_column_width(ws_tech)

    # Збереження у bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Ім'я файлу
    unique_dates = {
        (kp.event_date.date() if isinstance(kp.event_date, datetime) else kp.event_date)
        for kp in kps
        if kp.event_date
    }
    
    type_suffix = "box" if calculation_type == "box" else "catering"
    
    if len(unique_dates) == 1:
        only_date = next(iter(unique_dates))
        filename = f"purchase_{type_suffix}_{only_date}.xlsx"
    else:
        filename = f"purchase_{type_suffix}.xlsx"

    return buffer.getvalue(), filename
