"""
Сервіс для генерації Excel-файлу закупки з КП.
Формат відповідає закупка.ods з трьома типами аркушів:
1. "список продуктов" - зведений список продуктів з категоріями
2. "Меню {Назва КП}" - меню для кухарів БЕЗ ЦІН
3. "техкарта {Назва КП}" - техкарти страв з інгредієнтами
"""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, date, time
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Literal

from sqlalchemy.orm import Session, joinedload

import models
from recipe_service import get_recipe_by_name, get_recipe_by_id

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise RuntimeError("openpyxl is required for procurement Excel export") from e


# ============ Кольори та стилі ============

COLORS = {
    'yellow_header': 'FFFF00',      # Жовтий - категорії
    'light_yellow': 'FFFFCC',       # Світло-жовтий - підкатегорії  
    'orange_format': 'F79646',      # Помаранчевий - формат заходу
    'purple_dish': 'E6B8FF',        # Фіолетовий - страва в техкарті
    'white': 'FFFFFF',
    'light_gray': 'F2F2F2',
}

FONTS = {
    'header': Font(bold=True, size=12),
    'format': Font(bold=True, size=14),
    'category': Font(bold=True, size=11),
    'normal': Font(size=10),
}

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _safe_sheet_title(base: str) -> str:
    """Робить безпечну назву для аркуша Excel (макс. 31 символ)."""
    invalid_chars = set(r"[]:*?/\\")
    cleaned = "".join(c for c in base if c not in invalid_chars)
    if not cleaned:
        cleaned = "Sheet"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def _get_recipe_for_item(db: Session, item: models.Item) -> Optional[models.Recipe]:
    """Отримує техкарту для страви (по item_id або по назві)."""
    if item.id:
        # Шукаємо по item_id
        recipe = (
            db.query(models.Recipe)
            .options(
                joinedload(models.Recipe.ingredients),
                joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients)
            )
            .filter(models.Recipe.item_id == item.id)
            .first()
        )
        if recipe:
            return recipe
    
    # Fallback: по назві
    if item.name:
        return get_recipe_by_name(db, item.name)
    
    return None


def _parse_event_time(time_str: Optional[str]) -> Tuple[Optional[time], Optional[time]]:
    """Парсить час з рядка типу '15:00-21:30'."""
    if not time_str:
        return None, None
    
    try:
        parts = time_str.split('-')
        if len(parts) == 2:
            start = datetime.strptime(parts[0].strip(), '%H:%M').time()
            end = datetime.strptime(parts[1].strip(), '%H:%M').time()
            return start, end
    except:
        pass
    
    return None, None


def _format_date(dt: Optional[datetime]) -> str:
    """Форматує дату для Excel."""
    if not dt:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%d.%m.%Y")
    if isinstance(dt, date):
        return dt.strftime("%d.%m.%Y")
    return str(dt)


def _format_time(t: Optional[time]) -> str:
    """Форматує час для Excel."""
    if not t:
        return ""
    return t.strftime("%H:%M")


def _get_product_category(product_name: str, db: Session) -> Tuple[Optional[str], Optional[str]]:
    """Отримує категорію та підкатегорію продукту з таблиці products."""
    product_name_clean = product_name.strip()
    
    # Спочатку точний пошук
    product = (
        db.query(models.Product)
        .filter(models.Product.name.ilike(product_name_clean))
        .first()
    )
    if product:
        return product.category, product.subcategory
    
    # Потім пошук по частині назви
    product = (
        db.query(models.Product)
        .filter(models.Product.name.ilike(f"%{product_name_clean}%"))
        .first()
    )
    if product:
        return product.category, product.subcategory
    
    return None, None


def _collect_all_ingredients(
    db: Session,
    kps: List[models.KP]
) -> Dict[str, Dict]:
    """
    Збирає всі інгредієнти з усіх техкарт КП.
    Повертає: {product_name: {"total_grams": float, "category": str, "subcategory": str}}
    """
    ingredients_dict: Dict[str, Dict] = defaultdict(lambda: {"total_grams": 0.0, "category": None, "subcategory": None})
    
    # Завантажуємо всі техкарти
    item_ids = set()
    for kp in kps:
        for kp_item in kp.items:
            if kp_item.item_id:
                item_ids.add(kp_item.item_id)
    
    recipes_by_item_id: Dict[int, models.Recipe] = {}
    if item_ids:
        recipes = (
            db.query(models.Recipe)
            .options(
                joinedload(models.Recipe.ingredients),
                joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients)
            )
            .filter(models.Recipe.item_id.in_(list(item_ids)))
            .all()
        )
        for r in recipes:
            if r.item_id:
                recipes_by_item_id[int(r.item_id)] = r
    
    # Проходимо всі КП та страви
    for kp in kps:
        for kp_item in kp.items:
            quantity = kp_item.quantity or 0
            if quantity <= 0:
                continue
            
            recipe = None
            if kp_item.item_id:
                recipe = recipes_by_item_id.get(int(kp_item.item_id))
            
            if not recipe and kp_item.item:
                recipe = _get_recipe_for_item(db, kp_item.item)
            
            # Fallback: пошук по назві (для custom items)
            if not recipe:
                item_name = ""
                if kp_item.item and kp_item.item.name:
                    item_name = kp_item.item.name
                elif kp_item.name:
                    item_name = kp_item.name
                
                if item_name:
                    recipe = get_recipe_by_name(db, item_name)
            
            if not recipe:
                continue
            
            # Обчислюємо інгредієнти
            if recipe.recipe_type == "box" and recipe.components:
                # Бокси: вага × кількість_компонентів × кількість_порцій
                for component in recipe.components:
                    component_qty = component.quantity_per_portion or 1.0
                    for ingredient in component.ingredients:
                        total_weight = ingredient.weight_per_unit * component_qty * quantity
                        product_name = ingredient.product_name.strip()
                        ingredients_dict[product_name]["total_grams"] += total_weight
            else:
                # Кейтерінг: вага × кількість_порцій
                for ingredient in recipe.ingredients:
                    total_weight = ingredient.weight_per_portion * quantity
                    product_name = ingredient.product_name.strip()
                    ingredients_dict[product_name]["total_grams"] += total_weight
    
    # Отримуємо категорії для продуктів
    for product_name in ingredients_dict:
        category, subcategory = _get_product_category(product_name, db)
        ingredients_dict[product_name]["category"] = category
        ingredients_dict[product_name]["subcategory"] = subcategory
    
    return dict(ingredients_dict)


def _create_products_sheet(ws, ingredients_dict: Dict[str, Dict]):
    """Створює аркуш 'список продуктов' з категоріями."""
    ws.title = "список продуктов"
    
    # Заголовок
    ws['A1'] = "Позначення"
    ws['B1'] = "Закупка в грамах"
    ws['A1'].font = FONTS['header']
    ws['B1'].font = FONTS['header']
    ws['A1'].fill = PatternFill(start_color=COLORS['yellow_header'], end_color=COLORS['yellow_header'], fill_type="solid")
    ws['B1'].fill = PatternFill(start_color=COLORS['yellow_header'], end_color=COLORS['yellow_header'], fill_type="solid")
    ws['A1'].border = THIN_BORDER
    ws['B1'].border = THIN_BORDER
    
    row = 2
    
    # Групуємо по категоріях
    categories_dict: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))
    uncategorized = []
    
    for product_name, data in ingredients_dict.items():
        category = data.get("category") or "ІНШЕ"
        subcategory = data.get("subcategory") or ""
        categories_dict[category][subcategory].append((product_name, data["total_grams"]))
    
    # Сортуємо категорії
    sorted_categories = sorted(categories_dict.keys())
    
    for category in sorted_categories:
        # Заголовок категорії
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = FONTS['category']
        ws[f'A{row}'].fill = PatternFill(start_color=COLORS['yellow_header'], end_color=COLORS['yellow_header'], fill_type="solid")
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'B{row}'] = ""
        ws[f'B{row}'].border = THIN_BORDER
        row += 1
        
        # Підкатегорії
        subcategories = sorted(categories_dict[category].keys())
        for subcategory in subcategories:
            if subcategory:
                ws[f'A{row}'] = subcategory
                ws[f'A{row}'].font = FONTS['normal']
                ws[f'A{row}'].fill = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type="solid")
                ws[f'A{row}'].border = THIN_BORDER
                ws[f'B{row}'] = ""
                ws[f'B{row}'].border = THIN_BORDER
                row += 1
            
            # Продукти
            products = sorted(categories_dict[category][subcategory], key=lambda x: x[0])
            for product_name, total_grams in products:
                ws[f'A{row}'] = product_name
                ws[f'B{row}'] = round(total_grams, 2)
                ws[f'B{row}'].number_format = '0.00'
                ws[f'A{row}'].border = THIN_BORDER
                ws[f'B{row}'].border = THIN_BORDER
                row += 1
    
    # Налаштування ширини колонок
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 20


def _create_menu_sheet(ws, kp: models.KP, db: Session):
    """Створює аркуш 'Меню {Назва КП}' БЕЗ ЦІН."""
    safe_title = _safe_sheet_title(f"Меню {kp.title}")
    ws.title = safe_title
    
    row = 1
    
    # Шапка
    event_date = _format_date(kp.event_date)
    customer_name = kp.client_name or (kp.client.name if kp.client else "")
    location = kp.event_location or ""
    event_time_str = kp.event_time or ""
    
    # Парсимо час
    time_start, time_end = _parse_event_time(kp.event_time)
    time_range = ""
    if time_start and time_end:
        time_range = f"{_format_time(time_start)}-{_format_time(time_end)}"
    elif kp.event_time:
        time_range = kp.event_time
    
    ws[f'A{row}'] = "Дата"
    ws[f'C{row}'] = event_date
    ws[f'D{row}'] = time_range
    ws[f'A{row}'].font = FONTS['header']
    row += 1
    
    ws[f'A{row}'] = "Замовник"
    ws[f'B{row}'] = customer_name
    row += 1
    
    ws[f'A{row}'] = "Локація"
    ws[f'B{row}'] = location
    row += 1
    
    row += 1  # Порожній рядок
    
    ws[f'A{row}'] = "Коментар"
    comment = kp.booking_terms or ""
    ws[f'B{row}'] = comment
    row += 1
    
    row += 1  # Порожній рядок
    
    # Формати заходу
    event_formats = sorted(kp.event_formats, key=lambda x: x.order_index or 0) if kp.event_formats else []
    has_formats = len(event_formats) > 0
    
    if not has_formats:
        # Якщо немає форматів, показуємо всі страви в одному форматі
        format_name = kp.event_format or 'Загальне меню'
        format_time = kp.event_time or ""
        people_count = kp.people_count or 0
        
        # Заголовок формату
        format_text = format_name
        if format_time:
            format_text += f"({format_time})"
        
        ws[f'A{row}'] = "№ п/п"
        ws[f'B{row}'] = format_text
        if people_count:
            ws[f'B{row}'] = f"{format_text}\nМеню із розрахунку на {people_count} осіб"
        ws[f'C{row}'] = "Вихід на стіл"
        ws[f'D{row}'] = "Кіл-сть порцій"
        
        # Форматування заголовка формату
        ws[f'B{row}'].font = FONTS['format']
        ws[f'B{row}'].fill = PatternFill(start_color=COLORS['orange_format'], end_color=COLORS['orange_format'], fill_type="solid")
        ws[f'A{row}'].font = FONTS['header']
        ws[f'C{row}'].font = FONTS['header']
        ws[f'D{row}'].font = FONTS['header']
        ws[f'A{row}'].border = THIN_BORDER
        ws[f'B{row}'].border = THIN_BORDER
        ws[f'C{row}'].border = THIN_BORDER
        ws[f'D{row}'].border = THIN_BORDER
        row += 1
        
        # Групуємо всі страви по категоріях
        items_by_category: Dict[str, List[models.KPItem]] = defaultdict(list)
        for kp_item in kp.items:
            category = "Інше"
            if kp_item.item and kp_item.item.subcategory:
                category = kp_item.item.subcategory.name or "Інше"
            items_by_category[category].append(kp_item)
        
        # Сортуємо категорії
        sorted_categories = sorted(items_by_category.keys())
        
        for category in sorted_categories:
            # Заголовок категорії
            ws[f'B{row}'] = category
            ws[f'B{row}'].font = FONTS['category']
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['yellow_header'], end_color=COLORS['yellow_header'], fill_type="solid")
            ws[f'B{row}'].border = THIN_BORDER
            row += 1
            
            # Страви
            item_num = 1
            for kp_item in items_by_category[category]:
                item_name = ""
                if kp_item.item and kp_item.item.name:
                    item_name = kp_item.item.name
                elif kp_item.name:
                    item_name = kp_item.name
                else:
                    item_name = "Без назви"
                
                # Вихід на стіл (вага)
                weight_str = ""
                if kp_item.item and kp_item.item.weight:
                    weight_str = str(kp_item.item.weight)
                elif kp_item.weight:
                    weight_str = str(kp_item.weight)
                
                quantity = kp_item.quantity or 0
                
                ws[f'A{row}'] = item_num
                ws[f'B{row}'] = item_name
                ws[f'C{row}'] = weight_str
                ws[f'D{row}'] = quantity
                
                # Чергування кольорів
                if row % 2 == 0:
                    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                    ws[f'B{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                    ws[f'C{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                    ws[f'D{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                
                ws[f'A{row}'].border = THIN_BORDER
                ws[f'B{row}'].border = THIN_BORDER
                ws[f'C{row}'].border = THIN_BORDER
                ws[f'D{row}'].border = THIN_BORDER
                
                item_num += 1
                row += 1
            
            row += 1  # Порожній рядок після категорії
    
    else:
        # Є формати заходу
        for event_format in event_formats:
            format_name = event_format.name or ""
            format_time = event_format.event_time or ""
            people_count = event_format.people_count or kp.people_count or 0
            
            # Заголовок формату
            format_text = format_name
            if format_time:
                format_text += f"({format_time})"
            
            ws[f'A{row}'] = "№ п/п"
            ws[f'B{row}'] = format_text
            if people_count:
                ws[f'B{row}'] = f"{format_text}\nМеню із розрахунку на {people_count} осіб"
            ws[f'C{row}'] = "Вихід на стіл"
            ws[f'D{row}'] = "Кіл-сть порцій"
            
            # Форматування заголовка формату
            ws[f'B{row}'].font = FONTS['format']
            ws[f'B{row}'].fill = PatternFill(start_color=COLORS['orange_format'], end_color=COLORS['orange_format'], fill_type="solid")
            ws[f'A{row}'].font = FONTS['header']
            ws[f'C{row}'].font = FONTS['header']
            ws[f'D{row}'].font = FONTS['header']
            ws[f'A{row}'].border = THIN_BORDER
            ws[f'B{row}'].border = THIN_BORDER
            ws[f'C{row}'].border = THIN_BORDER
            ws[f'D{row}'].border = THIN_BORDER
            row += 1
            
            # Групуємо страви по категоріях
            items_by_category: Dict[str, List[models.KPItem]] = defaultdict(list)
            for kp_item in kp.items:
                if kp_item.event_format_id == event_format.id:
                    category = "Інше"
                    if kp_item.item and kp_item.item.subcategory:
                        category = kp_item.item.subcategory.name or "Інше"
                    items_by_category[category].append(kp_item)
            
            # Сортуємо категорії
            sorted_categories = sorted(items_by_category.keys())
            
            for category in sorted_categories:
                # Заголовок категорії
                ws[f'B{row}'] = category
                ws[f'B{row}'].font = FONTS['category']
                ws[f'B{row}'].fill = PatternFill(start_color=COLORS['yellow_header'], end_color=COLORS['yellow_header'], fill_type="solid")
                ws[f'B{row}'].border = THIN_BORDER
                row += 1
                
                # Страви
                item_num = 1
                for kp_item in items_by_category[category]:
                    item_name = ""
                    if kp_item.item and kp_item.item.name:
                        item_name = kp_item.item.name
                    elif kp_item.name:
                        item_name = kp_item.name
                    else:
                        item_name = "Без назви"
                    
                    # Вихід на стіл (вага)
                    weight_str = ""
                    if kp_item.item and kp_item.item.weight:
                        weight_str = str(kp_item.item.weight)
                    elif kp_item.weight:
                        weight_str = str(kp_item.weight)
                    
                    quantity = kp_item.quantity or 0
                    
                    ws[f'A{row}'] = item_num
                    ws[f'B{row}'] = item_name
                    ws[f'C{row}'] = weight_str
                    ws[f'D{row}'] = quantity
                    
                    # Чергування кольорів
                    if row % 2 == 0:
                        ws[f'A{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                        ws[f'B{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                        ws[f'C{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                        ws[f'D{row}'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type="solid")
                    
                    ws[f'A{row}'].border = THIN_BORDER
                    ws[f'B{row}'].border = THIN_BORDER
                    ws[f'C{row}'].border = THIN_BORDER
                    ws[f'D{row}'].border = THIN_BORDER
                    
                    item_num += 1
                    row += 1
                
                row += 1  # Порожній рядок після категорії
            
            row += 1  # Порожній рядок після формату
    
    # Налаштування ширини колонок
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20


def _create_recipe_sheet(ws, kp: models.KP, db: Session):
    """Створює аркуш 'техкарта {Назва КП}' з детальними техкартами."""
    safe_title = _safe_sheet_title(f"техкарта {kp.title}")
    ws.title = safe_title
    
    # Заголовки колонок
    ws['A1'] = "A (кількість_закупки)"
    ws['B1'] = "B (порцій)"
    ws['C1'] = "C (назва страви)"
    ws['D1'] = "D (інгредієнт)"
    ws['E1'] = "E (грам на порцію)"
    
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = FONTS['header']
        cell.border = THIN_BORDER
    
    row = 2
    
    # Завантажуємо техкарти
    item_ids = {kp_item.item_id for kp_item in kp.items if kp_item.item_id}
    recipes_by_item_id: Dict[int, models.Recipe] = {}
    
    if item_ids:
        recipes = (
            db.query(models.Recipe)
            .options(
                joinedload(models.Recipe.ingredients),
                joinedload(models.Recipe.components).joinedload(models.RecipeComponent.ingredients)
            )
            .filter(models.Recipe.item_id.in_(list(item_ids)))
            .all()
        )
        for r in recipes:
            if r.item_id:
                recipes_by_item_id[int(r.item_id)] = r
    
    # Проходимо всі страви в КП
    for kp_item in kp.items:
        quantity = kp_item.quantity or 0
        if quantity <= 0:
            continue
        
        item_name = ""
        if kp_item.item and kp_item.item.name:
            item_name = kp_item.item.name
        elif kp_item.name:
            item_name = kp_item.name
        else:
            item_name = "Без назви"
        
        recipe = None
        if kp_item.item_id:
            recipe = recipes_by_item_id.get(int(kp_item.item_id))
        
        if not recipe and kp_item.item:
            recipe = _get_recipe_for_item(db, kp_item.item)
        
        # Fallback: пошук по назві (для custom items)
        if not recipe and item_name and item_name != "Без назви":
            recipe = get_recipe_by_name(db, item_name)
        
        if not recipe:
            # Страва без техкарти - все одно додаємо рядок
            ws[f'A{row}'] = 0
            ws[f'B{row}'] = ""
            ws[f'C{row}'] = item_name
            ws[f'D{row}'] = ""
            ws[f'E{row}'] = ""
            ws[f'C{row}'].fill = PatternFill(start_color=COLORS['purple_dish'], end_color=COLORS['purple_dish'], fill_type="solid")
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1
            continue
        
        # Вихід на порцію
        portion_weight = recipe.weight_per_portion or 0
        portion_str = ""
        if portion_weight:
            portion_str = str(int(portion_weight))
        
        # Рядок страви
        ws[f'A{row}'] = ""  # Буде заповнено після підрахунку
        ws[f'B{row}'] = portion_str
        ws[f'C{row}'] = item_name
        ws[f'D{row}'] = ""
        ws[f'E{row}'] = ""
        ws[f'C{row}'].fill = PatternFill(start_color=COLORS['purple_dish'], end_color=COLORS['purple_dish'], fill_type="solid")
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        dish_row = row
        row += 1
        
        # Інгредієнти
        if recipe.recipe_type == "box" and recipe.components:
            # Бокси: компоненти → інгредієнти
            for component in recipe.components:
                component_qty = component.quantity_per_portion or 1.0
                for ingredient in component.ingredients:
                    grams_per_portion = ingredient.weight_per_unit * component_qty
                    total_grams = grams_per_portion * quantity
                    
                    ws[f'A{row}'] = round(total_grams, 2)
                    ws[f'B{row}'] = ""
                    ws[f'C{row}'] = ""
                    ws[f'D{row}'] = ingredient.product_name
                    ws[f'E{row}'] = round(grams_per_portion, 6)
                    ws[f'E{row}'].number_format = '0.000000'
                    
                    ws[f'D{row}'].fill = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type="solid")
                    for col in range(1, 6):
                        ws.cell(row=row, column=col).border = THIN_BORDER
                    row += 1
        else:
            # Кейтерінг: прямі інгредієнти
            for ingredient in recipe.ingredients:
                grams_per_portion = ingredient.weight_per_portion
                total_grams = grams_per_portion * quantity
                
                ws[f'A{row}'] = round(total_grams, 2)
                ws[f'B{row}'] = ""
                ws[f'C{row}'] = ""
                ws[f'D{row}'] = ingredient.product_name
                ws[f'E{row}'] = round(grams_per_portion, 6)
                ws[f'E{row}'].number_format = '0.000000'
                
                ws[f'D{row}'].fill = PatternFill(start_color=COLORS['light_yellow'], end_color=COLORS['light_yellow'], fill_type="solid")
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                row += 1
        
        # Підраховуємо загальну кількість для страви
        total_for_dish = 0
        for r in range(dish_row + 1, row):
            cell_value = ws[f'A{r}'].value
            if isinstance(cell_value, (int, float)):
                total_for_dish += cell_value
        ws[f'A{dish_row}'] = round(total_for_dish, 2)
    
    # Налаштування ширини колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20


def generate_procurement_excel(
    db: Session,
    kp_ids: List[int]
) -> Tuple[bytes, str]:
    """
    Генерує Excel-файл закупки на основі вибраних КП.
    
    Args:
        db: Сесія бази даних
        kp_ids: Список ID КП
    
    Повертає (bytes, filename).
    """
    if not kp_ids:
        raise ValueError("Список KP ID порожній")
    
    # Завантажуємо КП з усіма необхідними даними
    kps = (
        db.query(models.KP)
        .options(
            joinedload(models.KP.items).joinedload(models.KPItem.item).joinedload(models.Item.subcategory),
            joinedload(models.KP.event_formats),
            joinedload(models.KP.client)
        )
        .filter(models.KP.id.in_(kp_ids))
        .all()
    )
    
    if not kps:
        raise ValueError("Не знайдено жодного КП для вказаних ID")
    
    # Створюємо workbook
    wb = Workbook()
    wb.remove(wb.active)  # Видаляємо дефолтний аркуш
    
    # 1. Аркуш "список продуктов"
    all_ingredients = _collect_all_ingredients(db, kps)
    ws_products = wb.create_sheet()
    _create_products_sheet(ws_products, all_ingredients)
    
    # 2. Аркуші "Меню {Назва КП}" для кожної КП
    for kp in kps:
        ws_menu = wb.create_sheet()
        _create_menu_sheet(ws_menu, kp, db)
    
    # 3. Аркуші "техкарта {Назва КП}" для кожної КП
    for kp in kps:
        ws_recipe = wb.create_sheet()
        _create_recipe_sheet(ws_recipe, kp, db)
    
    # Збереження у bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Ім'я файлу
    unique_dates = {
        (kp.event_date.date() if isinstance(kp.event_date, datetime) else kp.event_date)
        for kp in kps
        if kp.event_date
    }
    
    if len(unique_dates) == 1:
        only_date = next(iter(unique_dates))
        if isinstance(only_date, date):
            filename = f"Закупка_{only_date.strftime('%d-%m-%Y')}.xlsx"
        else:
            filename = f"Закупка_{str(only_date)}.xlsx"
    else:
        filename = f"Закупка_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    
    return buffer.getvalue(), filename

